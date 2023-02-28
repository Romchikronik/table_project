import xlwt

# import xlsxwriter
# from djqscsv import write_csv, render_to_csv_response
from datetime import datetime, timedelta
from django.utils import timezone

from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout, get_user
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, response
from django.shortcuts import render, redirect

from .forms import *
from .utils import *


def loginPage(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        # try:
        #     User.objects.get(usernama=username)
        # except:
        #     messages.error(request, 'Попробуйте снова!')
        user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('main')
        else:
            messages.error(request, 'Имя пользователя или пароль не существует')
    context = {'title': 'Авторизация'}
    return render(request, 'mainapp/login.html', context)


@login_required
def mainPage(request):
    context = {'title': 'Главная старница'}
    return render(request, 'mainapp/index.html', context)


# def departament_view(request):
#     user = User
# request.user.has_perm('mainapp.departament_group'):
#     if user.groups.filter(name='department_group').exists():

# if user in

# @login_required
# def get_department(request, del_group_id, page_title, department_name, department_url):
#     if del_group_id(request):
#         # raise Exception('У вас нету доступа к этому департаменту')
#         messages.error(request, 'У вас нету доступа к этому департаменту')
#         return redirect('/')
#     else:
#         context = {
#             'title': page_title,
#             f'{department_name}': department_name
#         }
#         return render(request, f'mainapp/departments/{department_url}.html', context)


# @login_required
# def get_department_projects(request):
#     return get_department(request, del_group_loiha_id, 'Отдел - Лоиха', projects_department, 'projects_department')

@login_required
def get_department_projects(request):
    if del_group_loiha_id(request):
        # raise Exception('У вас нету доступа к этому департаменту')
        messages.error(request, 'У вас нету доступа к этому департаменту')
        return redirect('/')
    else:
        context = {
            'title': 'Отдел - Лоиха',
            'projects_department': projects_department
        }
        return render(request, 'mainapp/departments/projects_department.html', context)


@login_required
def get_department_exports(request):
    if del_group_export_id(request):
        # raise Exception('У вас нету доступа к этому департаменту')
        messages.error(request, 'У вас нету доступа к этому департаменту')
        return redirect('/')
    else:
        context = {
            'title': 'Отдел экспорта',  # Отдел - 2
            'second_department_tables_menu': second_department_tables_menu
        }
        return render(request, 'mainapp/departments/exports_department.html', context)


@login_required
def get_department_3(request):
    if del_group_vault_id(request):
        # raise Exception('У вас нету доступа к этому департаменту')
        messages.error(request, 'У вас нету доступа к этому департаменту')
        return redirect('/')
    else:
        context = {
            'title': 'Сводный отдел',  # Отдел - 3
            'third_department_tables_menu': third_department_tables_menu
        }
        return render(request, 'mainapp/departments/department_3.html', context)


@login_required
def get_department_4(request):
    if del_group_post_monitoring_id(request):
        messages.error(request, 'У вас нету доступа к этому департаменту')
        return redirect('/')
    else:
        context = {
            'title': 'Отдел постмониторинга',
            'fourth_department_tables_menu': fourth_department_tables_menu
        }
        return render(request, 'mainapp/departments/department_4.html', context)


@login_required
def get_department_5(request):
    context = {
        'title': 'Отдел - 5',
        'fifth_department_tables_menu': fifth_department_tables_menu
    }
    return render(request, 'mainapp/departments/department_5.html', context)


def export_excel_loiha41(request, filter_slug):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xls'
    # response['Content-Disposition'] = 'attachment; filename=table' + \
    #                                   str(datetime.now()) + '.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('table', cell_overwrite_ok=True)
    row_num = 2  # с какой строки начинается наша таблица
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    cell_style = xlwt.easyxf("font: bold on; align: vert center, horiz center")
    cell_title = xlwt.easyxf("font: bold on, height 280; align: vert center, horiz left")
    # cell_style = xlwt.easyxf("align: vert centre, horiz center")

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_loiha_id(request):
        ws.write_merge(0, 0, 0, 25, f'{request.user.district}, Илова-4.1', cell_title)
        ws.write_merge(1, 1, 0, 3, 'SANOAT', cell_style)
        ws.write_merge(1, 1, 4, 7, 'QISHLOQ', cell_style)
        ws.write_merge(1, 1, 8, 11, 'QURILISH', cell_style)
        ws.write_merge(1, 1, 12, 15, 'XIZMATLAR', cell_style)
        ws.write_merge(1, 1, 16, 19, 'CHAKANA SAVDO', cell_style)
        ws.write_merge(1, 1, 20, 21, 'TASHQI SAVDO AYLANMASI', cell_style)
        ws.write_merge(1, 1, 22, 23, 'EKSPORT', cell_style)
        ws.write_merge(1, 1, 24, 25, 'IMPORT', cell_style)
    else:
        ws.write_merge(0, 0, 0, 26, 'Илова-4.1', cell_title)
        # ws.write_merge(1, 2, 0, 1, 'Район', cell_style)
        ws.write_merge(1, 1, 1, 4, 'SANOAT', cell_style)
        ws.write_merge(1, 1, 5, 8, 'QISHLOQ', cell_style)
        ws.write_merge(1, 1, 9, 12, 'QURILISH', cell_style)
        ws.write_merge(1, 1, 13, 16, 'XIZMATLAR', cell_style)
        ws.write_merge(1, 1, 17, 20, 'CHAKANA SAVDO', cell_style)
        ws.write_merge(1, 1, 21, 22, 'TASHQI SAVDO AYLANMASI', cell_style)
        ws.write_merge(1, 1, 23, 24, 'EKSPORT', cell_style)
        ws.write_merge(1, 1, 25, 26, 'IMPORT', cell_style)

    # ws.col(0).width = 4500
    # ws.col(21).width = 5000

    # "Район",
    columns_list = ["mlrd. so'm",
                    "o'shish sur'ti %",
                    "prognoz",
                    "farqi (-;+)",
                    "mlrd. so'm",
                    "o'shish sur'ti %",
                    "prognoz",
                    "farqi (-;+)",
                    "mlrd. so'm",
                    "o'shish sur'ti%",
                    "prognoz",
                    "farqi (-;+)",
                    "mlrd. so'm",
                    "o'shish sur'ti%",
                    "prognoz",
                    "farqi (-;+)",
                    "mlrd. so'm",
                    "o'shish sur'ti%",
                    "prognoz",
                    "farqi (-;+)",
                    "ming. AQSh doll",
                    "o'shish sur'ti%",
                    "ming. AQSh doll",
                    "o'shish sur'ti%",
                    "ming. AQSh doll",
                    "o'shish sur'ti%"
                    ]

    if not get_group_loiha_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        ws.col(col_num).width = 3800
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    # date_style = xlwt.XFStyle()
    # time_create = datetime.strftime('time_create', '%d/%m/%y %h:%m:%s')
    fields = [
        # 'district__district',
        'bill_sum_industry',
        'picture_of_growth_industry',
        'forecast_industry',
        'difference_industry',
        'bill_sum_locality',
        'picture_of_growth_locality',
        'forecast_locality',
        'difference_locality',
        'bill_sum_construction',
        'picture_of_growth_construction',
        'forecast_construction',
        'difference_construction',
        'bill_sum_services',
        'picture_of_growth_services',
        'forecast_services',
        'difference_services',
        'bill_sum_retail',
        'picture_of_growth_retail',
        'forecast_retail',
        'difference_retail',
        'thousand_dollar_international_trade',
        'picture_of_growth_international_trade',
        'thousand_dollar_export',
        'picture_of_growth_export',
        'thousand_dollar_import',
        'picture_of_growth_import',
        # 'time_create'
    ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_loiha_tables(request, filter_slug, Loiha41, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)

    wb.save(response)
    return response


def export_excel_loiha52(request, filter_slug):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('table', cell_overwrite_ok=True)
    row_num = 2  # с какой строки начинается наша таблица
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    cell_style = xlwt.easyxf("font: bold on; align: vert center, horiz center")
    cell_title = xlwt.easyxf("font: bold on, height 280; align: vert center, horiz left")
    # cell_style = xlwt.easyxf("align: vert centre, horiz center")

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_loiha_id(request):
        ws.write_merge(0, 0, 0, 12, f'{request.user.district}, Илова-5.2', cell_title)
        ws.write_merge(1, 1, 4, 5, 'Состояние запасов на 01.01.2019', cell_style)
        ws.write_merge(1, 1, 6, 7, 'Оперативная информация', cell_style)
    else:
        ws.write_merge(0, 0, 0, 13, 'Илова-5.2', cell_title)
        # ws.write_merge(1, 2, 0, 1, 'Район', cell_style)
        ws.write_merge(1, 1, 5, 6, 'Состояние запасов на 01.01.2019', cell_style)
        ws.write_merge(1, 1, 7, 8, 'Оперативная информация', cell_style)

    # ws.col(0).width = 4500
    # ws.col(21).width = 5000

    # "Район",
    columns_list = ["Название места, где расположены природные ресурсы(Объекты геологоразведки)",
                    "Участки геологических образований",
                    "Наименование месторождений геологических формаций (сырых)",
                    "Единица измерения",
                    "A+B+C1 категории",
                    "С2 категория",
                    "Уровень развития",
                    "Дата и номер лицензии",
                    "Объем добычи в 2017 г.",
                    "Соответствующее министерство и ведомство",
                    "Подтвержденный отчет о запасах и его дата",
                    "Комментарий",
                    ]

    if not get_group_loiha_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        ws.col(col_num).width = 3800
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    # date_style = xlwt.XFStyle()
    # time_create = datetime.strftime('time_create', '%d/%m/%y %h:%m:%s')
    fields = [
        # 'district__district',
        'places_of_exploration',
        'sites_of_geo_formations',
        'name_of_formations',
        'unit',
        'abc_categories',
        'c2_category',
        'lvl_development_operational_information',
        'date_license_operational_information',
        'volume_of_production_2017',
        'ministry_and_department',
        'confirmed_stock_and_date',
        'comment'
        # 'time_create'
    ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_loiha_tables(request, filter_slug, Loiha52, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)

    wb.save(response)
    return response


def export_excel_loiha14(request, filter_slug):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('table', cell_overwrite_ok=True)
    row_num = 3  # с какой строки начинается наша таблица
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    cell_style = xlwt.easyxf("font: bold on; align: vert center, horiz center")
    cell_title = xlwt.easyxf("font: bold on, height 280; align: vert center, horiz left")
    # cell_style = xlwt.easyxf("align: vert centre, horiz center")

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_loiha_id(request):
        ws.write_merge(0, 0, 0, 13, f'{request.user.district}, Илова-14', cell_title)
        ws.write_merge(1, 2, 0, 2, 'Доступная информация', cell_style)
        ws.write_merge(1, 1, 3, 9, 'Процесс реализации запланированных проектов', cell_style)
        ws.write_merge(1, 1, 10, 12, 'Финансирование', cell_style)

        ws.write_merge(2, 2, 5, 7, 'Из них', cell_style)
        ws.write_merge(2, 2, 11, 12, 'Иностранные', cell_style)
    else:
        ws.write_merge(0, 0, 0, 14, 'Илова-14', cell_title)
        # ws.write_merge(1, 2, 0, 1, 'Район', cell_style)
        ws.write_merge(1, 2, 1, 3, 'Доступная информация', cell_style)
        ws.write_merge(1, 1, 4, 10, 'Процесс реализации запланированных проектов', cell_style)
        ws.write_merge(1, 1, 11, 13, 'Финансирование', cell_style)

        ws.write_merge(2, 2, 6, 8, 'Из них', cell_style)
        ws.write_merge(2, 2, 12, 13, 'Иностранные', cell_style)

    # ws.col(0).width = 4500
    # ws.col(21).width = 5000

    # "Район",
    columns_list = ["Организация",
                    "Род деятельности",
                    "Имеющиеся рабочие места",
                    "Проектная деятельность",
                    "Оценка (млрд долл)",
                    "Собственные средства",
                    "Банковский кредит",
                    "Иностранные инвестиции",
                    "Созданные рабочие места",
                    "Обьем экспорта (млрд долл)",
                    "Название банка",
                    "Название страны",
                    "Название организации",
                    ]

    if not get_group_loiha_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        ws.col(col_num).width = 3800
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    # date_style = xlwt.XFStyle()
    # time_create = datetime.strftime('time_create', '%d/%m/%y %h:%m:%s')
    fields = [
        'organization_available_information',
        'kind_of_activity_available_information',
        'available_jobs_available_information',
        'project_activity_at_work',
        'grade_at_work',
        'own_funds_grade_at_work',
        'bank_loan_grade_at_work',
        'sources_sources_of_financing',
        'created_new_jobs_at_work',
        'export_volume_at_work',
        'name_bank_financing',
        'name_country_financing',
        'organization_name_foreign_financing'
    ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_loiha_tables(request, filter_slug, Loiha14, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)

    wb.save(response)
    return response


def export_excel_loiha131(request, filter_slug):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('table', cell_overwrite_ok=True)
    row_num = 2  # с какой строки начинается наша таблица
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    cell_style = xlwt.easyxf("font: bold on; align: vert center, horiz center")
    cell_title = xlwt.easyxf("font: bold on, height 280; align: vert center, horiz left")
    # cell_style = xlwt.easyxf("align: vert centre, horiz center")

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_loiha_id(request):
        ws.write_merge(0, 0, 0, 10, f'{request.user.district}, Илова-13.1', cell_title)
        ws.write_merge(1, 1, 1, 8, 'Из них', cell_style)
    else:
        ws.write_merge(0, 0, 0, 11, 'Илова-13.1', cell_title)
        ws.write_merge(1, 1, 2, 9, 'Из них', cell_style)

    # "Район",
    columns_list = [
        "Известные проблемы",
        "Банк",
        "Земля и здание",
        "Коммунальные",
        "Таможня",
        "Налог",
        "Разрешения",
        "Бюрокартический барьер",
        "Другое",
        "Решенные проблемы"
    ]

    if not get_group_loiha_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        ws.col(col_num).width = 3800
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    # date_style = xlwt.XFStyle()
    # time_create = datetime.strftime('time_create', '%d/%m/%y %h:%m:%s')
    fields = [
        'identified_problems',
        'bank_identified_problems',
        'land_and_building_identified_problems',
        'public_services_identified_problems',
        'customs_identified_problems',
        'tax_identified_problems',
        'permission_identified_problems',
        'bureaucratic_obstacles_identified_problems',
        'other_identified_problems',
        'resolved_problems_identified_problems',
    ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_loiha_tables(request, filter_slug, Loiha131, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)

    wb.save(response)
    return response


def export_excel_loiha122(request, filter_slug):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('table', cell_overwrite_ok=True)
    row_num = 3  # с какой строки начинается наша таблица
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    cell_style = xlwt.easyxf("font: bold on; align: vert center, horiz center")
    cell_title = xlwt.easyxf("font: bold on, height 280; align: vert center, horiz left")
    # cell_style = xlwt.easyxf("align: vert centre, horiz center")

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_loiha_id(request):
        ws.write_merge(0, 0, 0, 12, f'{request.user.district}, Илова-12.2', cell_title)
        ws.write_merge(1, 1, 1, 8, 'в процессе', cell_style)
        ws.write_merge(1, 1, 9, 11, 'Финансирование', cell_style)

        ws.write_merge(2, 2, 4, 6, 'Из них', cell_style)
        ws.write_merge(2, 2, 10, 11, 'Иностранные', cell_style)
    else:
        ws.write_merge(0, 0, 0, 13, 'Илова-12.2', cell_title)
        ws.write_merge(1, 1, 2, 9, 'в процессе', cell_style)
        ws.write_merge(1, 1, 10, 12, 'Финансирование', cell_style)

        ws.write_merge(2, 2, 5, 7, 'Из них', cell_style)
        ws.write_merge(2, 2, 11, 12, 'Иностранные', cell_style)

    # "Район",
    columns_list = [
        "Сформированные советы",
        "Проекты",
        "Производсвенные мощности",
        "Оценка",
        "Собственные средства",
        "Банковский кредит",
        "Иностранные инвестиции",
        "Уровнь работы",
        "Обьем экспорта",
        "Название банка",
        "Название страны",
        "название организации",
    ]

    if not get_group_loiha_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        ws.col(col_num).width = 3800
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    # date_style = xlwt.XFStyle()
    # time_create = datetime.strftime('time_create', '%d/%m/%y %h:%m:%s')
    fields = [
        'formed_tips',
        'projects_at_work',
        'production_capacity_at_work',
        'grade_at_work',
        'own_funds_grade_at_work',
        'bank_loan_grade_at_work',
        'foreign_investments_grade_at_work',
        'workplace_at_work',
        'export_volume_at_work',
        'name_bank_financing',
        'name_country_financing',
        'organization_name_foreign_financing'
    ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_loiha_tables(request, filter_slug, Loiha122, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)

    wb.save(response)
    return response


def export_excel_loiha121(request, filter_slug):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('table', cell_overwrite_ok=True)
    row_num = 2  # с какой строки начинается наша таблица
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    cell_style = xlwt.easyxf("font: bold on; align: vert center, horiz center")
    cell_title = xlwt.easyxf("font: bold on, height 280; align: vert center, horiz left")
    # cell_style = xlwt.easyxf("align: vert centre, horiz center")

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_loiha_id(request):
        ws.write_merge(0, 0, 0, 11, f'{request.user.district}, Илова-12.1', cell_title)
        ws.write_merge(1, 1, 2, 3, 'Приоритетные цели', cell_style)
        ws.write_merge(1, 1, 4, 10, 'Заявленные результаты', cell_style)
    else:
        ws.write_merge(0, 0, 0, 12, 'Илова-12.1', cell_title)
        ws.write_merge(1, 1, 3, 4, 'Приоритетные цели', cell_style)
        ws.write_merge(1, 1, 5, 11, 'Заявленные результаты', cell_style)
    # "Район",
    columns_list = [
        "Утвержденные холдинговые компании количество",
        "Предложения предпринимателей(ф.и.о)",
        "Название организации",
        "Тип деятельности",
        "Проекты",
        "Производственная мощность",
        "Оценка",
        "Рабочее место",
        "Экспорт",
        "Другое",
        "Финансирование",
    ]

    if not get_group_loiha_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        ws.col(col_num).width = 3800
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    # date_style = xlwt.XFStyle()
    # time_create = datetime.strftime('time_create', '%d/%m/%y %h:%m:%s')
    fields = [
        'approved_holding_companies',
        'entrepreneur',
        'names_of_organizations_priority',
        'type_of_activity_priority',
        'projects_the_stated_results',
        'productive_capacity_stated_results',
        'grade_stated_results',
        'workplace_stated_results',
        'export_stated_results',
        'other_stated_results',
        'financing',
    ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_loiha_tables(request, filter_slug, Loiha121, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)

    wb.save(response)
    return response


def export_excel_loiha12(request, filter_slug):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('table', cell_overwrite_ok=True)
    row_num = 3  # с какой строки начинается наша таблица
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    cell_style = xlwt.easyxf("font: bold on; align: vert center, horiz center")
    cell_title = xlwt.easyxf("font: bold on, height 280; align: vert center, horiz left")
    # cell_style = xlwt.easyxf("align: vert centre, horiz center")

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_loiha_id(request):
        ws.write_merge(0, 0, 0, 12, f'{request.user.district}, Илова-12', cell_title)
        ws.write_merge(1, 1, 3, 11, 'Из них', cell_style)

        ws.write_merge(2, 2, 3, 5, 'Выбранные проекты', cell_style)
        ws.write_merge(2, 2, 6, 8, 'В процессе', cell_style)
        ws.write_merge(2, 2, 9, 11, 'Будут завершены к концу года', cell_style)
    else:
        ws.write_merge(0, 0, 0, 13, 'Илова-12', cell_title)
        ws.write_merge(1, 1, 4, 12, 'Из них', cell_style)

        ws.write_merge(2, 2, 4, 6, 'Выбранные проекты', cell_style)
        ws.write_merge(2, 2, 7, 9, 'В процессе', cell_style)
        ws.write_merge(2, 2, 10, 12, 'Будут завершены к концу года', cell_style)
    # "Район",
    columns_list = [
        "Все проекты",
        "Оценка(млрд долл)",
        "Рабочее место",
        "Проекты",
        "Оценка",
        "Рабочее место",
        "Проекты",
        "Оценка (млрд долл)",
        "Рабочее место",
        "Проекты",
        "Оценка (млрд долл)",
        "Рабочее место",
    ]

    if not get_group_loiha_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        ws.col(col_num).width = 3800
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    # date_style = xlwt.XFStyle()
    # time_create = datetime.strftime('time_create', '%d/%m/%y %h:%m:%s')
    fields = [
        'all_projects',
        'grade',
        'workplace',
        'projects_selected_projects_from_all_projects',
        'grade_selected_projects_from_all_projects',
        'workplace_selected_projects_from_all_projects',
        'all_projects_at_work_from_all_projects',
        'grade_at_work_from_all_projects',
        'workplace_at_work_from_all_projects',
        'all_projects_will_be_completed_end_the_year_from_all_projects',
        'grade_will_be_completed_end_the_year_all_projects',
        'workplace_will_be_completed_end_the_year_all_projects',
    ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_loiha_tables(request, filter_slug, Loiha12, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)

    wb.save(response)
    return response


def export_excel_loiha10(request, filter_slug):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('table', cell_overwrite_ok=True)
    row_num = 2  # с какой строки начинается наша таблица
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    cell_style = xlwt.easyxf("font: bold on; align: vert center, horiz center")
    cell_title = xlwt.easyxf("font: bold on, height 280; align: vert center, horiz left")
    # cell_style = xlwt.easyxf("align: vert centre, horiz center")

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_loiha_id(request):
        ws.write_merge(0, 0, 0, 20, f'{request.user.district}, Илова-10', cell_title)
        ws.write_merge(1, 1, 2, 3, 'Из них', cell_style)

        ws.write_merge(1, 1, 5, 16, 'Специализация', cell_style)
    else:
        ws.write_merge(0, 0, 0, 21, 'Илова-10', cell_title)
        ws.write_merge(1, 1, 3, 4, 'Из них', cell_style)
        ws.write_merge(1, 1, 6, 17, 'Специализация', cell_style)
    # "Район",
    columns_list = [
        "насленный пункт (количество)",
        "Жилые дома (количество)",
        "Желающие заняться предпринимательством (количество)",
        "процентное соотношение",
        "дома занятые в секторе",
        "животноводство",
        "птицеводство",
        "кроликоведение",
        "пчеловодство",
        "фермерство",
        "садоводство",
        "парники",
        "преподование",
        "мелькое производство",
        "туризм",
        "услуги",
        "другие",
        "железная тетрадь оставшиеся",
        "молодёжная тетрадь оставшиеся",
        "женская тетрадь оставшиеся",
    ]

    if not get_group_loiha_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        ws.col(col_num).width = 3800
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    # date_style = xlwt.XFStyle()
    # time_create = datetime.strftime('time_create', '%d/%m/%y %h:%m:%s')
    fields = [
        'locality_amount',
        'residential_buildings_amount',
        'who_want_to_do_business_residential_buildings',
        'percent_residential_buildings',
        'employed_in_the_sector',
        'animal_husbandry_speciality',
        'poultry_farming_specialty',
        'rabbing_breeding_specialty',
        'Beekeeping_specialty',
        'farm_specialty',
        'gardening_specialty',
        'greenhouses_specialty',
        'teaching_specialty',
        'small_production_specialty',
        'tourism_specialty',
        'services_specialty',
        'other_specialty',
        'remaining_on_the_iron_notebook',
        'remaining_on_the_youth_notebook',
        'remaining_on_the_womens_notebook',
    ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_loiha_tables(request, filter_slug, Loiha10, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)

    wb.save(response)
    return response


def export_excel_loiha6(request, filter_slug):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('table', cell_overwrite_ok=True)
    row_num = 2  # с какой строки начинается наша таблица
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    cell_style = xlwt.easyxf("font: bold on; align: vert center, horiz center")
    cell_title = xlwt.easyxf("font: bold on, height 280; align: vert center, horiz left")
    # cell_style = xlwt.easyxf("align: vert centre, horiz center")

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_loiha_id(request):
        ws.write_merge(0, 0, 0, 11, f'{request.user.district}, Илова-6', cell_title)
        ws.write_merge(1, 1, 1, 4, 'Из них', cell_style)
        ws.write_merge(1, 1, 6, 7, 'Из них', cell_style)
        ws.write_merge(1, 1, 9, 10, 'Из них', cell_style)
    else:
        ws.write_merge(0, 0, 0, 12, 'Илова-6', cell_title)
        ws.write_merge(1, 1, 2, 5, 'Из них', cell_style)
        ws.write_merge(1, 1, 7, 8, 'Из них', cell_style)
        ws.write_merge(1, 1, 10, 11, 'Из них', cell_style)

    # "Район",
    columns_list = [
        "Пустующие объекты",
        "адрес",
        "владелец",
        "государственное имущество (количество)",
        "частное имущуство (количество)",
        "Неиспользуемые производственные площадки(Га)",
        "ориентированное на сельское хозяйство(Га)",
        "ориентированное на производство(Га)",
        "Предложения по инвестициям (источники)",
        "Предварительная стоимость проекта(млн.долл)",
        "количество созданных навых рабочих мест",
    ]

    if not get_group_loiha_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        ws.col(col_num).width = 3800
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    # date_style = xlwt.XFStyle()
    # time_create = datetime.strftime('time_create', '%d/%m/%y %h:%m:%s')
    fields = [
        'empty_objects',
        'adress_empty_objects',
        'owner_empty_objects',
        'state_property_amount_empty_objects',
        'private_property_amount_empty_objects',
        'unused_production_sites',
        'focused_on_agriculture_unused_production_sites',
        'focused_on_production_unused_production_sites',
        'Investment_proposals_sources',
        'preliminary_project_cost_bil_sum_sources',
        'new_jobs_created_amount_sources'
    ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_loiha_tables(request, filter_slug, Loiha6, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)

    wb.save(response)
    return response


def export_excel_loiha13(request, filter_slug):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('table', cell_overwrite_ok=True)
    row_num = 2  # с какой строки начинается наша таблица
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    cell_style = xlwt.easyxf("font: bold on; align: vert center, horiz center")
    cell_title = xlwt.easyxf("font: bold on, height 280; align: vert center, horiz left")
    # cell_style = xlwt.easyxf("align: vert centre, horiz center")

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_loiha_id(request):
        ws.write_merge(0, 0, 0, 7, f'{request.user.district}, Илова-13', cell_title)
        ws.write_merge(1, 1, 3, 4, 'Источники финансирования', cell_style)

    else:
        ws.write_merge(0, 0, 0, 8, 'Илова-13', cell_title)
        ws.write_merge(1, 1, 4, 5, 'Источники финансирования', cell_style)

    # "Район",
    columns_list = [
        "Название проекта",
        "Проектная деятельность",
        "Проектная мощность",
        "Оценка в миллиардах долларах",
        "Иностранные инвестиции",
        "Cозданные новые рабочие места",
        "Источники",
    ]

    if not get_group_loiha_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        ws.col(col_num).width = 3800
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    # date_style = xlwt.XFStyle()
    # time_create = datetime.strftime('time_create', '%d/%m/%y %h:%m:%s')
    fields = [
        'project_name',
        'project_activity',
        'project_capacity',
        'grate_sources_of_financing',
        'foreign_financing_grate_sources_of_financing',
        'sources_sources_of_financing',
        'created_new_jobs'
    ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_loiha_tables(request, filter_slug, Loiha13, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)

    wb.save(response)
    return response


# export second department
def export_excel_sanoat(request, filter_slug):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('table', cell_overwrite_ok=True)
    row_num = 1  # с какой строки начинается наша таблица
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    cell_style = xlwt.easyxf("font: bold on; align: vert center, horiz center")
    cell_title = xlwt.easyxf("font: bold on, height 280; align: vert center, horiz left")
    # cell_style = xlwt.easyxf("align: vert centre, horiz center")

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_export_id(request):
        ws.write_merge(0, 0, 0, 18, f'{request.user.district}, Sanoat', cell_title)
    else:
        ws.write_merge(0, 0, 0, 19, 'Sanoat', cell_title)

    # ws.col(0).width = 4500
    # ws.col(21).width = 5000

    # "Район",
    columns_list = [
        "ИНН отправ./получ.",
        "ИНН",
        "Наим. отправителя",
        "Адрес отправителя",
        "Наим. получателя",
        "Адрес получателя",
        "Инн лица отв. за финансовое урегулирование",
        "лицо отв. за финансовое урегулирование",
        "Адрес лица отв. за финансовое урегулирование",
        "Валюта контракта",
        "Фактурная стоимость",
        "Код товара",
        "Наим. товара",
        "Вес. нетто",
        "Стат. стоимость",
        "Номер и дата контракта",
        "ИДН",
        "Страна отправления/назначения",
        "Дата выпуска",
    ]

    if not get_group_export_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        ws.col(col_num).width = 3800
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    # date_style = xlwt.XFStyle()
    # time_create = datetime.strftime('time_create', '%d/%m/%y %h:%m:%s')
    fields = [
        'inn_of_sender_or_recipient',
        'inn',
        'name_of_sender',
        'address_of_sender',
        'name_of_recipient',
        'address_of_recipient',
        'financial_responsible_inn',
        'face_responsible_for_finance',
        'address_of_face_responsible_for_finance',
        'currency_of_contract',
        'invoice_value',
        'code_of_goods',
        'name_of_goods',
        'weight_netto',
        'stat_price',
        'number_and_date_of_contract',
        'idn',
        'destination_country',
        'date_of_issue'
    ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_export_tables(request, filter_slug, Sanoat, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)

    wb.save(response)
    return response


def export_excel_kx(request, filter_slug):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('table', cell_overwrite_ok=True)
    row_num = 1  # с какой строки начинается наша таблица
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    cell_style = xlwt.easyxf("font: bold on; align: vert center, horiz center")
    cell_title = xlwt.easyxf("font: bold on, height 280; align: vert center, horiz left")
    # cell_style = xlwt.easyxf("align: vert centre, horiz center")

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_export_id(request):
        ws.write_merge(0, 0, 0, 19, f'{request.user.district}, KX', cell_title)
    else:
        ws.write_merge(0, 0, 0, 20, 'KX', cell_title)

    # ws.col(0).width = 4500
    # ws.col(21).width = 5000

    # "Район",
    columns_list = [
        "ИНН отправ./получ.",
        "ИНН",
        "Наим. отправителя",
        "Адрес отправителя",
        "Наим. получателя",
        "Адрес получателя",
        "Инн лица отв. за финансовое урегулирование",
        "лицо отв. за финансовое урегулирование",
        "Адрес лица отв. за финансовое урегулирование",
        "Валюта контракта",
        "Фактурная стоимость",
        "Код товара",
        "Наим. товара",
        "Вес. нетто",
        "Тонна",
        "Стат. стоимость",
        "Номер и дата контракта",
        "ИДН",
        "Страна отправления/назначения",
        "Дата выпуска",
    ]

    if not get_group_export_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        ws.col(col_num).width = 3800
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    # date_style = xlwt.XFStyle()
    # time_create = datetime.strftime('time_create', '%d/%m/%y %h:%m:%s')
    fields = [
        'inn_of_sender_or_recipient',
        'inn',
        'name_of_sender',
        'address_of_sender',
        'name_of_recipient',
        'address_of_recipient',
        'financial_responsible_inn',
        'face_responsible_for_finance',
        'address_of_face_responsible_for_finance',
        'currency_of_contract',
        'invoice_value',
        'code_of_goods',
        'name_of_goods',
        'weight_netto',
        'ton',
        'stat_price',
        'number_and_date_of_contract',
        'idn',
        'destination_country',
        'date_of_issue'
    ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_export_tables(request, filter_slug, KH, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)

    wb.save(response)
    return response


def export_excel_table1(request, filter_slug):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('table', cell_overwrite_ok=True)
    row_num = 3  # с какой строки начинается наша таблица
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    cell_style = xlwt.easyxf("font: bold on; align: vert center, horiz center")
    cell_title = xlwt.easyxf("font: bold on, height 280; align: vert center, horiz left")
    # cell_style = xlwt.easyxf("align: vert centre, horiz center")

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_export_id(request):
        ws.write_merge(0, 0, 0, 45, f'{request.user.district}, Таблица 1', cell_title)
        ws.write_merge(1, 1, 5, 8, 'Всего', cell_style)
        ws.write_merge(1, 1, 9, 11, 'Январь-март', cell_style)
        ws.write_merge(1, 1, 12, 14, 'Январь', cell_style)
        ws.write_merge(1, 1, 15, 17, 'Февраль', cell_style)
        ws.write_merge(1, 1, 18, 20, 'Март', cell_style)
        ws.write_merge(1, 1, 21, 23, 'Январь-Апрель', cell_style)
        ws.write_merge(1, 1, 24, 26, 'Январь-Май', cell_style)
        ws.write_merge(1, 1, 27, 29, 'Январь-Июнь', cell_style)
        ws.write_merge(1, 1, 30, 32, 'Январь-Июль', cell_style)
        ws.write_merge(1, 1, 33, 35, 'Январь-Август', cell_style)
        ws.write_merge(1, 1, 36, 38, 'Январь-Сентябрь', cell_style)
        ws.write_merge(1, 1, 39, 41, '24-окт', cell_style)
        ws.write_merge(1, 1, 42, 44, 'Октябрь', cell_style)

        ws.write_merge(2, 2, 7, 8, 'плод', cell_style)
        ws.write_merge(2, 2, 10, 11, 'плод', cell_style)
        ws.write_merge(2, 2, 13, 14, 'плод', cell_style)
        ws.write_merge(2, 2, 16, 17, 'плод', cell_style)
        ws.write_merge(2, 2, 19, 20, 'плод', cell_style)
        ws.write_merge(2, 2, 22, 23, 'плод', cell_style)
        ws.write_merge(2, 2, 25, 26, 'плод', cell_style)
        ws.write_merge(2, 2, 28, 29, 'плод', cell_style)
        ws.write_merge(2, 2, 31, 32, 'плод', cell_style)
        ws.write_merge(2, 2, 34, 35, 'плод', cell_style)
        ws.write_merge(2, 2, 37, 38, 'плод', cell_style)
        ws.write_merge(2, 2, 40, 41, 'плод', cell_style)
        ws.write_merge(2, 2, 43, 44, 'плод', cell_style)

    else:
        ws.write_merge(0, 0, 0, 46, 'Таблица 1', cell_title)
        ws.write_merge(1, 1, 6, 9, 'Всего', cell_style)
        ws.write_merge(1, 1, 10, 12, 'Январь-март', cell_style)
        ws.write_merge(1, 1, 13, 15, 'Январь', cell_style)
        ws.write_merge(1, 1, 16, 18, 'Февраль', cell_style)
        ws.write_merge(1, 1, 19, 21, 'Март', cell_style)

        ws.write_merge(1, 1, 22, 24, 'Январь-Апрель', cell_style)
        ws.write_merge(1, 1, 25, 27, 'Январь-Май', cell_style)
        ws.write_merge(1, 1, 28, 30, 'Январь-Июнь', cell_style)
        ws.write_merge(1, 1, 31, 33, 'Январь-Июль', cell_style)
        ws.write_merge(1, 1, 34, 36, 'Январь-Август', cell_style)
        ws.write_merge(1, 1, 37, 39, 'Январь-Сентябрь', cell_style)
        ws.write_merge(1, 1, 40, 42, '24-окт', cell_style)
        ws.write_merge(1, 1, 43, 45, 'Октябрь', cell_style)

        ws.write_merge(2, 2, 8, 9, 'плод', cell_style)
        ws.write_merge(2, 2, 11, 12, 'плод', cell_style)
        ws.write_merge(2, 2, 14, 15, 'плод', cell_style)
        ws.write_merge(2, 2, 17, 18, 'плод', cell_style)
        ws.write_merge(2, 2, 20, 21, 'плод', cell_style)
        ws.write_merge(2, 2, 23, 24, 'плод', cell_style)
        ws.write_merge(2, 2, 26, 27, 'плод', cell_style)
        ws.write_merge(2, 2, 29, 30, 'плод', cell_style)
        ws.write_merge(2, 2, 32, 33, 'плод', cell_style)
        ws.write_merge(2, 2, 35, 36, 'плод', cell_style)
        ws.write_merge(2, 2, 38, 39, 'плод', cell_style)
        ws.write_merge(2, 2, 41, 42, 'плод', cell_style)
        ws.write_merge(2, 2, 44, 45, 'плод', cell_style)

    # ws.col(0).width = 4500
    # ws.col(21).width = 5000

    # "Район",
    columns_list = [
        "",
        "",
        "Область",
        "",
        "Наименование организаций в разрезе регионов республики",
        "Всего",
        "пром",
        "тонна",
        "сумма",
        "пром",
        "тонна",
        "сумма",
        "пром",
        "тонна",
        "сумма",
        "пром",
        "тонна",
        "сумма",
        "пром",
        "тонна",
        "сумма",
        "пром",
        "тонна",
        "сумма",
        "пром",
        "тонна",
        "сумма",
        "пром",
        "тонна",
        "сумма",
        "пром",
        "тонна",
        "сумма",
        "пром",
        "тонна",
        "сумма",
        "пром",
        "тонна",
        "сумма",
        "пром",
        "тонна",
        "сумма",
        "пром",
        "тонна",
        "сумма",
        "Ички экспорт",
    ]

    if not get_group_export_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        ws.col(col_num).width = 3800
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    # date_style = xlwt.XFStyle()
    # time_create = datetime.strftime('time_create', '%d/%m/%y %h:%m:%s')
    fields = [
        'table_id',
        'cell_name',
        'region',
        'date',
        'total',
        'total_total',
        'total_prom',
        'total_ton',
        'total_sum',
        'january_march_prom',
        'january_march_ton',
        'january_march_sum',
        'january_prom',
        'january_ton',
        'january_sum',
        'february_prom',
        'february_ton',
        'february_sum',
        'march_prom',
        'march_ton',
        'march_sum',
        'january_april_prom',
        'january_april_ton',
        'january_april_sum',
        'january_may_prom',
        'january_may_ton',
        'january_may_sum',
        'january_june_prom',
        'january_june_ton',
        'january_june_sum',
        'january_july_prom',
        'january_july_ton',
        'january_july_sum',
        'january_august_prom',
        'january_august_ton',
        'january_august_sum',
        'january_september_prom',
        'january_september_ton',
        'january_september_sum',
        'october_24_prom',
        'october_24_ton',
        'october_24_sum',
        'october_prom',
        'october_ton',
        'october_sum',
        'export',
    ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_export_tables(request, filter_slug, FirstTable, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)

    wb.save(response)
    return response


def export_excel_kunliu(request, filter_slug):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('table', cell_overwrite_ok=True)
    row_num = 1  # с какой строки начинается наша таблица
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    cell_style = xlwt.easyxf("font: bold on; align: vert center, horiz center")
    cell_title = xlwt.easyxf("font: bold on, height 280; align: vert center, horiz left")
    # cell_style = xlwt.easyxf("align: vert centre, horiz center")

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_export_id(request):
        ws.write_merge(0, 0, 0, 3, f'{request.user.district}, кунлиу', cell_title)
    else:
        ws.write_merge(0, 0, 0, 4, 'кунлиу', cell_title)

    # ws.col(0).width = 4500
    # ws.col(21).width = 5000

    # "Район",
    columns_list = [
        "Дата прогноза",
        "Всего",
        "саноат маҳсулотлари",
        "мева-сабзавотлар",
    ]

    if not get_group_export_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        ws.col(col_num).width = 3800
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    # date_style = xlwt.XFStyle()
    # time_create = datetime.strftime('time_create', '%d/%m/%y %h:%m:%s')
    fields = [
        'date_of_forecast',
        'overall',
        'sanoat',
        'meva_sabz',
    ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_export_tables(request, filter_slug, Kunliu, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)

    wb.save(response)
    return response

# Изменить поля всех остальных таблиц на *all_fields
# export third department


def export_excel_bank(request, filter_slug):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('table', cell_overwrite_ok=True)
    row_num = 3  # с какой строки начинается наша таблица
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    cell_style = xlwt.easyxf("font: bold on; align: vert center, horiz center")
    cell_title = xlwt.easyxf("font: bold on, height 280; align: vert center, horiz left")
    # cell_style = xlwt.easyxf("align: vert centre, horiz center")

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_vault_id(request):
        ws.write_merge(0, 0, 0, 22, f'{request.user.district}, Cвод банк', cell_title)
        ws.write_merge(1, 2, 1, 2, 'Лойиҳа сони', cell_style)
        ws.write_merge(1, 2, 3, 4, 'Умумий қиймати млн.сўм', cell_style)
        ws.write_merge(1, 1, 5, 12, 'шундан', cell_style)
        ws.write_merge(1, 2, 13, 14, 'Янги иш ўринлари сони', cell_style)
        ws.write_merge(1, 1, 15, 22, 'Иқтисодий самарадорлик', cell_style)

        ws.write_merge(2, 2, 5, 6, 'ўз маблағлари млн.сўм', cell_style)
        ws.write_merge(2, 2, 7, 8, 'банк кредитлари млн.сўм', cell_style)
        ws.write_merge(2, 2, 9, 10, 'хорижий кредитлар минг.долл', cell_style)
        ws.write_merge(2, 2, 11, 12, 'хорижий инвестициялар минг.долл', cell_style)
        ws.write_merge(2, 2, 15, 16, 'Ишлаб чиқариш қуввати млн.сўм', cell_style)
        ws.write_merge(2, 2, 17, 18, 'Импорт урнини босиш минг.долл', cell_style)
        ws.write_merge(2, 2, 19, 20, 'Экспорт ҳажми минг.долл', cell_style)
        ws.write_merge(2, 2, 21, 22, 'Бюджетга қўшимча тушум млн.сўм', cell_style)
    else:
        ws.write_merge(0, 0, 0, 23, 'Cвод банк', cell_title)
        ws.write_merge(1, 2, 2, 3, 'Лойиҳа сони', cell_style)
        ws.write_merge(1, 2, 4, 5, 'Умумий қиймати млн.сўм', cell_style)
        ws.write_merge(1, 1, 6, 13, 'шундан', cell_style)
        ws.write_merge(1, 2, 14, 15, 'Янги иш ўринлари сони', cell_style)
        ws.write_merge(1, 1, 16, 23, 'Иқтисодий самарадорлик', cell_style)

        ws.write_merge(2, 2, 6, 7, 'ўз маблағлари млн.сўм', cell_style)
        ws.write_merge(2, 2, 8, 9, 'банк кредитлари млн.сўм', cell_style)
        ws.write_merge(2, 2, 10, 11, 'хорижий кредитлар минг.долл', cell_style)
        ws.write_merge(2, 2, 12, 13, 'хорижий инвестициялар минг.долл', cell_style)
        ws.write_merge(2, 2, 16, 17, 'Ишлаб чиқариш қуввати млн.сўм', cell_style)
        ws.write_merge(2, 2, 18, 19, 'Импорт урнини босиш минг.долл', cell_style)
        ws.write_merge(2, 2, 20, 21, 'Экспорт ҳажми минг.долл', cell_style)
        ws.write_merge(2, 2, 22, 23, 'Бюджетга қўшимча тушум млн.сўм', cell_style)

    # ws.col(0).width = 4500
    # ws.col(21).width = 5000

    # "Район",
    columns_list = [
        "Banklar nomi",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
    ]

    if not get_group_vault_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        ws.col(col_num).width = 3800
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    # date_style = xlwt.XFStyle()
    # time_create = datetime.strftime('time_create', '%d/%m/%y %h:%m:%s')
    form = TableBankForm()
    all_fields = form.fields.keys()
    fields = [
        *all_fields
    ]
    # fieldsn = [
    #     'bank_name',
    #     'loiha_soni_reja',
    #     'loiha_soni_amalda',
    #     'umumiy_kiymati_reja',
    #     'umumiy_kiymati_amalda',
    #     'uz_mablag_reja',
    #     'uz_mablag_amalda',
    #     'bank_kredit_reja',
    #     'bank_kredit_amalda',
    #     'xorijiy_kredit_reja',
    #     'xorijiy_kredit_amalda',
    #     'xorijiy_invest_reja',
    #     'xorijiy_invest_amalda',
    #
    #     'yangi_ish_reja',
    #     'yangi_ish_amalda',
    #     'ishlab_chiqarish_reja',
    #     'ishlab_chiqarish_amalda',
    #     'import_reja',
    #     'import_amalda',
    #     'export_reja',
    #     'export_amalda',
    #     'budget_reja',
    #     'budget_amalda',
    # ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_vault_tables(request, filter_slug, BankVault, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)

    wb.save(response)
    return response


def export_excel_jami(request, filter_slug):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('table', cell_overwrite_ok=True)
    row_num = 3  # с какой строки начинается наша таблица
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    cell_style = xlwt.easyxf("font: bold on; align: vert center, horiz center")
    cell_title = xlwt.easyxf("font: bold on, height 280; align: vert center, horiz left")
    # cell_style = xlwt.easyxf("align: vert centre, horiz center")

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_vault_id(request):
        ws.write_merge(0, 0, 0, 21, f'{request.user.district}, Жами свод', cell_title)
        ws.write_merge(1, 2, 0, 1, 'Лойиҳа сони', cell_style)
        ws.write_merge(1, 2, 2, 3, 'Умумий қиймати млн.сўм', cell_style)
        ws.write_merge(1, 1, 4, 11, 'шундан', cell_style)
        ws.write_merge(1, 2, 12, 13, 'Янги иш ўринлари сони', cell_style)
        ws.write_merge(1, 1, 14, 21, 'Иқтисодий самарадорлик', cell_style)

        ws.write_merge(2, 2, 4, 5, 'ўз маблағлари млн.сўм', cell_style)
        ws.write_merge(2, 2, 6, 7, 'банк кредитлари млн.сўм', cell_style)
        ws.write_merge(2, 2, 8, 9, 'хорижий кредитлар минг.долл', cell_style)
        ws.write_merge(2, 2, 10, 11, 'хорижий инвестициялар минг.долл', cell_style)
        ws.write_merge(2, 2, 14, 15, 'Ишлаб чиқариш қуввати млн.сўм', cell_style)
        ws.write_merge(2, 2, 16, 17, 'Импорт урнини босиш минг.долл', cell_style)
        ws.write_merge(2, 2, 18, 19, 'Экспорт ҳажми минг.долл', cell_style)
        ws.write_merge(2, 2, 20, 21, 'Бюджетга қўшимча тушум млн.сўм', cell_style)
    else:
        ws.write_merge(0, 0, 0, 22, 'Жами свод', cell_title)
        ws.write_merge(1, 2, 1, 2, 'Лойиҳа сони', cell_style)
        ws.write_merge(1, 2, 3, 4, 'Умумий қиймати млн.сўм', cell_style)
        ws.write_merge(1, 1, 5, 12, 'шундан', cell_style)
        ws.write_merge(1, 2, 13, 14, 'Янги иш ўринлари сони', cell_style)
        ws.write_merge(1, 1, 15, 22, 'Иқтисодий самарадорлик', cell_style)

        ws.write_merge(2, 2, 5, 6, 'ўз маблағлари млн.сўм', cell_style)
        ws.write_merge(2, 2, 7, 8, 'банк кредитлари млн.сўм', cell_style)
        ws.write_merge(2, 2, 9, 10, 'хорижий кредитлар минг.долл', cell_style)
        ws.write_merge(2, 2, 11, 12, 'хорижий инвестициялар минг.долл', cell_style)
        ws.write_merge(2, 2, 15, 16, 'Ишлаб чиқариш қуввати млн.сўм', cell_style)
        ws.write_merge(2, 2, 17, 18, 'Импорт урнини босиш минг.долл', cell_style)
        ws.write_merge(2, 2, 19, 20, 'Экспорт ҳажми минг.долл', cell_style)
        ws.write_merge(2, 2, 21, 22, 'Бюджетга қўшимча тушум млн.сўм', cell_style)

    # ws.col(0).width = 4500
    # ws.col(21).width = 5000

    # "Район",
    columns_list = [
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
    ]

    if not get_group_vault_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        ws.col(col_num).width = 3800
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()

    form = TableJamiForm()
    all_fields = form.fields.keys()
    fields = [
        *all_fields
    ]

    # fields = [
    #     'loiha_soni_reja',
    #     'loiha_soni_amalda',
    #     'umumiy_kiymati_reja',
    #     'umumiy_kiymati_amalda',
    #     'uz_mablag_reja',
    #     'uz_mablag_amalda',
    #     'bank_kredit_reja',
    #     'bank_kredit_amalda',
    #     'xorijiy_kredit_reja',
    #     'xorijiy_kredit_amalda',
    #     'xorijiy_invest_reja',
    #     'xorijiy_invest_amalda',
    #     'yangi_ish_reja',
    #     'yangi_ish_amalda',
    #     'ishlab_chiqarish_reja',
    #     'ishlab_chiqarish_amalda',
    #     'import_reja',
    #     'import_amalda',
    #     'export_reja',
    #     'export_amalda',
    #     'budget_reja',
    #     'budget_amalda',
    # ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_vault_tables(request, filter_slug, JamiVault, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)

    wb.save(response)
    return response


def export_excel_reja(request, filter_slug):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('table', cell_overwrite_ok=True)
    row_num = 3  # с какой строки начинается наша таблица
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    cell_style = xlwt.easyxf("font: bold on; align: vert center, horiz center")
    cell_title = xlwt.easyxf("font: bold on, height 280; align: vert center, horiz left")
    # cell_style = xlwt.easyxf("align: vert centre, horiz center")

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_vault_id(request):
        ws.write_merge(0, 0, 0, 55, f'{request.user.district}, Cвод режа', cell_title)
        ws.write_merge(1, 2, 0, 6, 'Режа (ҳисобот даври)', cell_style)
        ws.write_merge(1, 2, 7, 13, 'Режага нисбатан ишга тушган', cell_style)
        ws.write_merge(1, 2, 14, 19, 'Ишга тушмаган', cell_style)
        ws.write_merge(1, 1, 20, 34, 'шундан', cell_style)
        ws.write_merge(1, 2, 35, 41, 'Муддатидан аввал ишга тушган лойиҳалар', cell_style)
        ws.write_merge(1, 2, 42, 48, 'Қўшимча (резерв) лойиҳалар', cell_style)
        ws.write_merge(1, 2, 49, 55, 'Жами ишга тушган лойиҳалар', cell_style)

        ws.write_merge(2, 2, 20, 26, 'Истиқболсиз', cell_style)
        ws.write_merge(2, 2, 27, 34, 'Муддатини узайтириш', cell_style)

    else:
        ws.write_merge(0, 0, 0, 56, 'Cвод режа', cell_title)

        ws.write_merge(1, 2, 1, 7, 'Режа (ҳисобот даври)', cell_style)
        ws.write_merge(1, 2, 8, 14, 'Режага нисбатан ишга тушган', cell_style)
        ws.write_merge(1, 2, 15, 20, 'Ишга тушмаган', cell_style)
        ws.write_merge(1, 1, 21, 35, 'шундан', cell_style)
        ws.write_merge(1, 2, 36, 42, 'Муддатидан аввал ишга тушган лойиҳалар', cell_style)
        ws.write_merge(1, 2, 43, 49, 'Қўшимча (резерв) лойиҳалар', cell_style)
        ws.write_merge(1, 2, 50, 56, 'Жами ишга тушган лойиҳалар', cell_style)

        ws.write_merge(2, 2, 21, 27, 'Истиқболсиз', cell_style)
        ws.write_merge(2, 2, 28, 35, 'Муддатини узайтириш', cell_style)

    # ws.col(0).width = 4500
    # ws.col(21).width = 5000

    # "Район",
    columns_list = [
        "Лойиҳа сони",
        "Лойиҳа қиймати, (млн.сўм)",
        "ўз маблағлари млн.сўм",
        "банк кредитлари млн.сўм",
        "хорижий кредитлар минг.долл",
        "хорижий инвестициялар минг.долл",
        "Иш ўрни",
        "Лойиҳа сони",
        "Лойиҳа қиймати, (млн.сўм)",
        "ўз маблағлари млн.сўм",
        "банк кредитлари млн.сўм",
        "хорижий кредитлар минг.долл",
        "хорижий инвестициялар минг.долл",
        "Иш ўрни",
        "Лойиҳа сони",
        "Лойиҳа қиймати, (млн.сўм)",
        "ўз маблағлари млн.сўм",
        "банк кредитлари млн.сўм",
        "хорижий кредитлар минг.долл",
        "хорижий инвестициялар минг.долл",
        "Иш ўрни",
        "Лойиҳа сони",
        "Лойиҳа қиймати, (млн.сўм)",
        "ўз маблағлари млн.сўм",
        "банк кредитлари млн.сўм",
        "хорижий кредитлар минг.долл",
        "хорижий инвестициялар минг.долл",
        "Иш ўрни",
        "Лойиҳа сони",
        "Лойиҳа қиймати, (млн.сўм)",
        "ўз маблағлари млн.сўм",
        "банк кредитлари млн.сўм",
        "хорижий кредитлар минг.долл",
        "хорижий инвестициялар минг.долл",
        "Иш ўрни",
        "Лойиҳа сони",
        "Лойиҳа қиймати, (млн.сўм)",
        "ўз маблағлари млн.сўм",
        "банк кредитлари млн.сўм",
        "хорижий кредитлар минг.долл",
        "хорижий инвестициялар минг.долл",
        "Иш ўрни",
        "Лойиҳа сони",
        "Лойиҳа қиймати, (млн.сўм)",
        "ўз маблағлари млн.сўм",
        "банк кредитлари млн.сўм",
        "хорижий кредитлар минг.долл",
        "хорижий инвестициялар минг.долл",
        "Иш ўрни",
        "Лойиҳа сони",
        "Лойиҳа қиймати, (млн.сўм)",
        "ўз маблағлари млн.сўм",
        "банк кредитлари млн.сўм",
        "хорижий кредитлар минг.долл",
        "хорижий инвестициялар минг.долл",
        "Иш ўрни",
    ]

    if not get_group_vault_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        ws.col(col_num).width = 3800
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()

    form = TableRejaForm()
    all_fields = form.fields.keys()
    fields = [
        *all_fields
    ]

    # fields = [
    #     'reja_loiha_soni',
    #     'reja_loiha_kiymati',
    #     'reja_mablag',
    #     'reja_kredit',
    #     'reja_xorijiy_kredit',
    #     'reja_xorijiy_invest',
    #     'reja_ish',
    #
    #     'nisbatan_loiha_soni',
    #     'nisbatan_loiha_kiymati',
    #     'nisbatan_mablag',
    #     'nisbatan_kredit',
    #     'nisbatan_xorijiy_kredit',
    #     'nisbatan_xorijiy_invest',
    #     'nisbatan_ish',
    #
    #     'ish_loiha_soni',
    #     'ish_loiha_kiymati',
    #     'ish_mablag',
    #     'ish_kredit',
    #     'ish_xorijiy_kredit',
    #     'ish_xorijiy_invest',
    #
    #     'istik_ish',
    #     'istik_loiha_soni',
    #     'istik_loiha_kiymati',
    #     'istik_mablag',
    #     'istik_kredit',
    #     'istik_xorijiy_kredit',
    #     'istik_xorijiy_invest',
    #
    #     'mud_ish',
    #     'mud_loiha_soni',
    #     'mud_loiha_kiymati',
    #     'mud_mablag',
    #     'mud_kredit',
    #     'mud_xorijiy_kredit',
    #     'mud_xorijiy_invest',
    #     'mud_ish_second',
    #
    #     'aval_loiha_soni',
    #     'aval_loiha_kiymati',
    #     'aval_mablag',
    #     'aval_kredit',
    #     'aval_xorijiy_kredit',
    #     'aval_xorijiy_invest',
    #     'aval_ish',
    #
    #     'reserve_loiha_soni',
    #     'reserve_loiha_kiymati',
    #     'reserve_mablag',
    #     'reserve_kredit',
    #     'reserve_xorijiy_kredit',
    #     'reserve_xorijiy_invest',
    #     'reserve_ish',
    #
    #     'total_loiha_soni',
    #     'total_loiha_kiymati',
    #     'total_mablag',
    #     'total_kredit',
    #     'total_xorijiy_kredit',
    #     'total_xorijiy_invest',
    #     'total_ish',
    # ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_vault_tables(request, filter_slug, RejaVault, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)

    wb.save(response)
    return response


# TODO Сделать поля Select
def export_excel_tarmok(request, filter_slug):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('table', cell_overwrite_ok=True)
    row_num = 3  # с какой строки начинается наша таблица
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    cell_style = xlwt.easyxf("font: bold on; align: vert center, horiz center")
    cell_title = xlwt.easyxf("font: bold on, height 280; align: vert center, horiz left")
    # cell_style = xlwt.easyxf("align: vert centre, horiz center")

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_vault_id(request):
        ws.write_merge(0, 0, 0, 23, f'{request.user.district}, Cвод тармок', cell_title)
        ws.write_merge(1, 2, 2, 3, 'Лойиҳа сони', cell_style)
        ws.write_merge(1, 2, 4, 5, 'Умумий қиймати млн.сўм', cell_style)
        ws.write_merge(1, 1, 6, 13, 'шундан', cell_style)
        ws.write_merge(1, 2, 14, 15, 'Янги иш ўринлари сони', cell_style)
        ws.write_merge(1, 1, 16, 23, 'Иқтисодий самарадорлик', cell_style)

        ws.write_merge(2, 2, 6, 7, 'ўз маблағлари млн.сўм', cell_style)
        ws.write_merge(2, 2, 8, 9, 'банк кредитлари млн.сўм', cell_style)
        ws.write_merge(2, 2, 10, 11, 'хорижий кредитлар минг.долл', cell_style)
        ws.write_merge(2, 2, 12, 13, 'хорижий инвестициялар минг.долл', cell_style)
        ws.write_merge(2, 2, 16, 17, 'Ишлаб чиқариш қуввати млн.сўм', cell_style)
        ws.write_merge(2, 2, 18, 19, 'Импорт урнини босиш минг.долл', cell_style)
        ws.write_merge(2, 2, 20, 21, 'Экспорт ҳажми минг.долл', cell_style)
        ws.write_merge(2, 2, 22, 23, 'Бюджетга қўшимча тушум млн.сўм', cell_style)
    else:
        ws.write_merge(0, 0, 0, 24, 'Cвод тармок', cell_title)
        ws.write_merge(1, 2, 3, 4, 'Лойиҳа сони', cell_style)
        ws.write_merge(1, 2, 5, 6, 'Умумий қиймати млн.сўм', cell_style)
        ws.write_merge(1, 1, 7, 14, 'шундан', cell_style)
        ws.write_merge(1, 2, 15, 16, 'Янги иш ўринлари сони', cell_style)
        ws.write_merge(1, 1, 17, 24, 'Иқтисодий самарадорлик', cell_style)

        ws.write_merge(2, 2, 7, 8, 'ўз маблағлари млн.сўм', cell_style)
        ws.write_merge(2, 2, 9, 10, 'банк кредитлари млн.сўм', cell_style)
        ws.write_merge(2, 2, 11, 12, 'хорижий кредитлар минг.долл', cell_style)
        ws.write_merge(2, 2, 13, 14, 'хорижий инвестициялар минг.долл', cell_style)
        ws.write_merge(2, 2, 17, 18, 'Ишлаб чиқариш қуввати млн.сўм', cell_style)
        ws.write_merge(2, 2, 19, 20, 'Импорт урнини босиш минг.долл', cell_style)
        ws.write_merge(2, 2, 21, 22, 'Экспорт ҳажми минг.долл', cell_style)
        ws.write_merge(2, 2, 23, 24, 'Бюджетга қўшимча тушум млн.сўм', cell_style)

    # ws.col(0).width = 4500
    # ws.col(21).width = 5000

    # "Район",
    columns_list = [
        "Категория",
        "Тармоқлар номи",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
        "Режа",
        "Амалда",
    ]

    if not get_group_vault_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        ws.col(col_num).width = 3800
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    # date_style = xlwt.XFStyle()
    # time_create = datetime.strftime('time_create', '%d/%m/%y %h:%m:%s')
    # tarmok_model = show_data_table(request, TarmokVault)
    #
    # fields_with_choices = [f.name for f in TarmokVault._meta.get_fields() if f.choices]
    #
    # # Create a new list of fields that includes the choice values
    # new_fields = []
    # for field_name in fields_with_choices:
    #     # Get the field instance
    #     field = TarmokVault._meta.get_field(field_name)
    #
    #     # Get the choices as a list of tuples
    #     choices = field.choices
    #
    #     # Create a new list of choice values
    #     choice_values = [choice[1] for choice in choices]
    #
    #     # Add the choice values to the field list
    #     new_fields.append(field_name)
    #     new_fields.extend(choice_values)

    form = TableTarmokForm()
    all_fields = form.fields.keys()
    fields = [
        *all_fields
    ]

    # for obj in all_fields:
    #     row = []
    #     for field in TarmokVault._meta.fields:
    #         if field.choices:
    #             row.append(getattr(obj, f'get_{field.name}_display'))
    #
    #     fields.insert(0, row)

    # fields = [
    #     *all_fields,
    # ]
    # fields = [
    #     'category',
    #     'industry',
    #
    #     'loiha_soni_reja',
    #     'loiha_soni_amalda',
    #     'umumiy_kiymati_reja',
    #     'umumiy_kiymati_amalda',
    #     'uz_mablag_reja',
    #     'uz_mablag_amalda',
    #     'bank_kredit_reja',
    #     'bank_kredit_amalda',
    #     'xorijiy_kredit_reja',
    #     'xorijiy_kredit_amalda',
    #     'xorijiy_invest_reja',
    #     'xorijiy_invest_amalda',
    #
    #     'yangi_ish_reja',
    #     'yangi_ish_amalda',
    #     'ishlab_chiqarish_reja',
    #     'ishlab_chiqarish_amalda',
    #     'import_reja',
    #     'import_amalda',
    #     'export_reja',
    #     'export_amalda',
    #     'budget_reja',
    #     'budget_amalda',
    # ]
    # fields.extend(new_fields)

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_vault_tables(request, filter_slug, TarmokVault, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)

    wb.save(response)
    return response


def export_excel_quarter(request, filter_slug):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('table', cell_overwrite_ok=True)
    row_num = 4  # с какой строки начинается наша таблица
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    cell_style = xlwt.easyxf("font: bold on; align: vert center, horiz center")
    cell_title = xlwt.easyxf("font: bold on, height 280; align: vert center, horiz left")
    # cell_style = xlwt.easyxf("align: vert centre, horiz center")

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_vault_id(request):
        ws.write_merge(0, 0, 0, 87, f'{request.user.district}, Cвод чорак', cell_title)
        ws.write_merge(1, 3, 0, 1, 'Лойиҳа сони', cell_style)
        ws.write_merge(1, 1, 2, 21, '1-Чорак', cell_style)

        ws.write_merge(1, 3, 22, 23, 'Лойиҳа сони', cell_style)
        ws.write_merge(1, 1, 24, 43, '2-Чорак', cell_style)

        ws.write_merge(1, 3, 44, 45, 'Лойиҳа сони', cell_style)
        ws.write_merge(1, 1, 46, 65, '3-Чорак', cell_style)

        ws.write_merge(1, 3, 66, 67, 'Лойиҳа сони', cell_style)
        ws.write_merge(1, 1, 68, 87, '4-Чорак', cell_style)

        ws.write_merge(2, 3, 2, 3, 'Умумий қиймати млн.сўм', cell_style)
        ws.write_merge(2, 2, 4, 11, 'шундан', cell_style)
        ws.write_merge(2, 3, 12, 13, 'Янги иш ўринлари сони', cell_style)
        ws.write_merge(2, 2, 14, 21, 'Иқтисодий самарадорлик', cell_style)

        ws.write_merge(3, 3, 4, 5, 'ўз маблағлари млн.сўм', cell_style)
        ws.write_merge(3, 3, 6, 7, 'банк кредитлари млн.сўм', cell_style)
        ws.write_merge(3, 3, 8, 9, 'хорижий кредитлар минг.долл', cell_style)
        ws.write_merge(3, 3, 10, 11, 'хорижий инвестициялар минг.долл', cell_style)
        ws.write_merge(3, 3, 14, 15, 'Ишлаб чиқариш қуввати млн.сўм', cell_style)
        ws.write_merge(3, 3, 16, 17, 'Импорт урнини босиш минг.долл', cell_style)
        ws.write_merge(3, 3, 18, 19, 'Экспорт ҳажми минг.долл', cell_style)
        ws.write_merge(3, 3, 20, 21, 'Бюджетга қўшимча тушум млн.сўм', cell_style)

        ws.write_merge(2, 3, 24, 25, 'Умумий қиймати млн.сўм', cell_style)
        ws.write_merge(2, 2, 26, 33, 'шундан', cell_style)
        ws.write_merge(2, 3, 34, 35, 'Янги иш ўринлари сони', cell_style)
        ws.write_merge(2, 2, 36, 43, 'Иқтисодий самарадорлик', cell_style)

        ws.write_merge(3, 3, 26, 27, 'ўз маблағлари млн.сўм', cell_style)
        ws.write_merge(3, 3, 28, 29, 'банк кредитлари млн.сўм', cell_style)
        ws.write_merge(3, 3, 30, 31, 'хорижий кредитлар минг.долл', cell_style)
        ws.write_merge(3, 3, 32, 33, 'хорижий инвестициялар минг.долл', cell_style)
        ws.write_merge(3, 3, 36, 37, 'Ишлаб чиқариш қуввати млн.сўм', cell_style)
        ws.write_merge(3, 3, 38, 39, 'Импорт урнини босиш минг.долл', cell_style)
        ws.write_merge(3, 3, 40, 41, 'Экспорт ҳажми минг.долл', cell_style)
        ws.write_merge(3, 3, 42, 43, 'Бюджетга қўшимча тушум млн.сўм', cell_style)

        ws.write_merge(2, 3, 46, 47, 'Умумий қиймати млн.сўм', cell_style)
        ws.write_merge(2, 2, 48, 55, 'шундан', cell_style)
        ws.write_merge(2, 3, 56, 57, 'Янги иш ўринлари сони', cell_style)
        ws.write_merge(2, 2, 58, 65, 'Иқтисодий самарадорлик', cell_style)

        ws.write_merge(3, 3, 48, 49, 'ўз маблағлари млн.сўм', cell_style)
        ws.write_merge(3, 3, 50, 51, 'банк кредитлари млн.сўм', cell_style)
        ws.write_merge(3, 3, 52, 53, 'хорижий кредитлар минг.долл', cell_style)
        ws.write_merge(3, 3, 54, 55, 'хорижий инвестициялар минг.долл', cell_style)
        ws.write_merge(3, 3, 58, 59, 'Ишлаб чиқариш қуввати млн.сўм', cell_style)
        ws.write_merge(3, 3, 60, 61, 'Импорт урнини босиш минг.долл', cell_style)
        ws.write_merge(3, 3, 62, 63, 'Экспорт ҳажми минг.долл', cell_style)
        ws.write_merge(3, 3, 64, 65, 'Бюджетга қўшимча тушум млн.сўм', cell_style)

        ws.write_merge(2, 3, 68, 69, 'Умумий қиймати млн.сўм', cell_style)
        ws.write_merge(2, 2, 70, 77, 'шундан', cell_style)
        ws.write_merge(2, 3, 78, 79, 'Янги иш ўринлари сони', cell_style)
        ws.write_merge(2, 2, 80, 87, 'Иқтисодий самарадорлик', cell_style)

        ws.write_merge(3, 3, 70, 71, 'ўз маблағлари млн.сўм', cell_style)
        ws.write_merge(3, 3, 72, 73, 'банк кредитлари млн.сўм', cell_style)
        ws.write_merge(3, 3, 74, 75, 'хорижий кредитлар минг.долл', cell_style)
        ws.write_merge(3, 3, 76, 77, 'хорижий инвестициялар минг.долл', cell_style)
        ws.write_merge(3, 3, 80, 81, 'Ишлаб чиқариш қуввати млн.сўм', cell_style)
        ws.write_merge(3, 3, 82, 83, 'Импорт урнини босиш минг.долл', cell_style)
        ws.write_merge(3, 3, 84, 85, 'Экспорт ҳажми минг.долл', cell_style)
        ws.write_merge(3, 3, 86, 87, 'Бюджетга қўшимча тушум млн.сўм', cell_style)

    else:
        ws.write_merge(0, 0, 0, 88, 'Cвод чорак', cell_title)

        ws.write_merge(1, 3, 1, 2, 'Лойиҳа сони', cell_style)
        ws.write_merge(1, 1, 3, 22, '1-Чорак', cell_style)

        ws.write_merge(1, 3, 23, 24, 'Лойиҳа сони', cell_style)
        ws.write_merge(1, 1, 25, 44, '2-Чорак', cell_style)

        ws.write_merge(1, 3, 45, 46, 'Лойиҳа сони', cell_style)
        ws.write_merge(1, 1, 47, 66, '3-Чорак', cell_style)

        ws.write_merge(1, 3, 67, 68, 'Лойиҳа сони', cell_style)
        ws.write_merge(1, 1, 69, 88, '4-Чорак', cell_style)

        ws.write_merge(2, 3, 3, 4, 'Умумий қиймати млн.сўм', cell_style)
        ws.write_merge(2, 2, 5, 12, 'шундан', cell_style)
        ws.write_merge(2, 3, 13, 14, 'Янги иш ўринлари сони', cell_style)
        ws.write_merge(2, 2, 15, 22, 'Иқтисодий самарадорлик', cell_style)

        ws.write_merge(3, 3, 5, 6, 'ўз маблағлари млн.сўм', cell_style)
        ws.write_merge(3, 3, 7, 8, 'банк кредитлари млн.сўм', cell_style)
        ws.write_merge(3, 3, 9, 10, 'хорижий кредитлар минг.долл', cell_style)
        ws.write_merge(3, 3, 11, 12, 'хорижий инвестициялар минг.долл', cell_style)
        ws.write_merge(3, 3, 15, 16, 'Ишлаб чиқариш қуввати млн.сўм', cell_style)
        ws.write_merge(3, 3, 17, 18, 'Импорт урнини босиш минг.долл', cell_style)
        ws.write_merge(3, 3, 19, 20, 'Экспорт ҳажми минг.долл', cell_style)
        ws.write_merge(3, 3, 21, 22, 'Бюджетга қўшимча тушум млн.сўм', cell_style)

        ws.write_merge(2, 3, 25, 26, 'Умумий қиймати млн.сўм', cell_style)
        ws.write_merge(2, 2, 27, 34, 'шундан', cell_style)
        ws.write_merge(2, 3, 35, 36, 'Янги иш ўринлари сони', cell_style)
        ws.write_merge(2, 2, 37, 44, 'Иқтисодий самарадорлик', cell_style)

        ws.write_merge(3, 3, 27, 28, 'ўз маблағлари млн.сўм', cell_style)
        ws.write_merge(3, 3, 29, 30, 'банк кредитлари млн.сўм', cell_style)
        ws.write_merge(3, 3, 31, 32, 'хорижий кредитлар минг.долл', cell_style)
        ws.write_merge(3, 3, 33, 34, 'хорижий инвестициялар минг.долл', cell_style)
        ws.write_merge(3, 3, 37, 38, 'Ишлаб чиқариш қуввати млн.сўм', cell_style)
        ws.write_merge(3, 3, 39, 40, 'Импорт урнини босиш минг.долл', cell_style)
        ws.write_merge(3, 3, 41, 42, 'Экспорт ҳажми минг.долл', cell_style)
        ws.write_merge(3, 3, 43, 44, 'Бюджетга қўшимча тушум млн.сўм', cell_style)

        ws.write_merge(2, 3, 47, 48, 'Умумий қиймати млн.сўм', cell_style)
        ws.write_merge(2, 2, 49, 56, 'шундан', cell_style)
        ws.write_merge(2, 3, 57, 58, 'Янги иш ўринлари сони', cell_style)
        ws.write_merge(2, 2, 59, 66, 'Иқтисодий самарадорлик', cell_style)

        ws.write_merge(3, 3, 49, 50, 'ўз маблағлари млн.сўм', cell_style)
        ws.write_merge(3, 3, 51, 52, 'банк кредитлари млн.сўм', cell_style)
        ws.write_merge(3, 3, 53, 54, 'хорижий кредитлар минг.долл', cell_style)
        ws.write_merge(3, 3, 55, 56, 'хорижий инвестициялар минг.долл', cell_style)
        ws.write_merge(3, 3, 59, 60, 'Ишлаб чиқариш қуввати млн.сўм', cell_style)
        ws.write_merge(3, 3, 61, 62, 'Импорт урнини босиш минг.долл', cell_style)
        ws.write_merge(3, 3, 63, 64, 'Экспорт ҳажми минг.долл', cell_style)
        ws.write_merge(3, 3, 65, 66, 'Бюджетга қўшимча тушум млн.сўм', cell_style)

        ws.write_merge(2, 3, 69, 70, 'Умумий қиймати млн.сўм', cell_style)
        ws.write_merge(2, 2, 71, 78, 'шундан', cell_style)
        ws.write_merge(2, 3, 79, 80, 'Янги иш ўринлари сони', cell_style)
        ws.write_merge(2, 2, 81, 88, 'Иқтисодий самарадорлик', cell_style)

        ws.write_merge(3, 3, 71, 72, 'ўз маблағлари млн.сўм', cell_style)
        ws.write_merge(3, 3, 73, 74, 'банк кредитлари млн.сўм', cell_style)
        ws.write_merge(3, 3, 75, 76, 'хорижий кредитлар минг.долл', cell_style)
        ws.write_merge(3, 3, 77, 78, 'хорижий инвестициялар минг.долл', cell_style)
        ws.write_merge(3, 3, 81, 82, 'Ишлаб чиқариш қуввати млн.сўм', cell_style)
        ws.write_merge(3, 3, 83, 84, 'Импорт урнини босиш минг.долл', cell_style)
        ws.write_merge(3, 3, 85, 86, 'Экспорт ҳажми минг.долл', cell_style)
        ws.write_merge(3, 3, 87, 88, 'Бюджетга қўшимча тушум млн.сўм', cell_style)

    # ws.col(0).width = 4500
    # ws.col(21).width = 5000

    # "Район",
    columns_list = [
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
    ]

    if not get_group_vault_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        ws.col(col_num).width = 3800
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    # date_style = xlwt.XFStyle()
    # time_create = datetime.strftime('time_create', '%d/%m/%y %h:%m:%s')
    form = TableQuarterForm()
    all_fields = form.fields.keys()
    fields = [
        *all_fields
    ]

    # fields = [
    #     'loiha_soni_reja',
    #     'loiha_soni_amalda',
    #     'umumiy_kiymati_reja',
    #     'umumiy_kiymati_amalda',
    #     'uz_mablag_reja',
    #     'uz_mablag_amalda',
    #     'bank_kredit_reja',
    #     'bank_kredit_amalda',
    #     'xorijiy_kredit_reja',
    #     'xorijiy_kredit_amalda',
    #     'xorijiy_invest_reja',
    #     'xorijiy_invest_amalda',
    #     'first_quarter_yangi_ish_reja',
    #     'first_quarter_yangi_ish_amalda',
    #     'first_quarter_ishlab_chiqarish_reja',
    #     'first_quarter_ishlab_chiqarish_amalda',
    #     'first_quarter_import_reja',
    #     'first_quarter_import_amalda',
    #     'first_quarter_export_reja',
    #     'first_quarter_export_amalda',
    #     'first_quarter_budget_reja',
    #     'first_quarter_budget_amalda',
    #
    #     'second_quarter_loiha_soni_reja',
    #     'second_quarter_loiha_soni_amalda',
    #     'second_quarter_umumiy_kiymati_reja',
    #     'second_quarter_umumiy_kiymati_amalda',
    #     'second_quarter_uz_mablag_reja',
    #     'second_quarter_uz_mablag_amalda',
    #     'second_quarter_bank_kredit_reja',
    #     'second_quarter_bank_kredit_amalda',
    #     'second_quarter_xorijiy_kredit_reja',
    #     'second_quarter_xorijiy_kredit_amalda',
    #     'second_quarter_xorijiy_invest_reja',
    #     'second_quarter_xorijiy_invest_amalda',
    #     'second_quarter_yangi_ish_reja',
    #     'second_quarter_yangi_ish_amalda',
    #     'second_quarter_ishlab_chiqarish_reja',
    #     'second_quarter_ishlab_chiqarish_amalda',
    #     'second_quarter_import_reja',
    #     'second_quarter_import_amalda',
    #     'second_quarter_export_reja',
    #     'second_quarter_export_amalda',
    #     'second_quarter_budget_reja',
    #     'second_quarter_budget_amalda',
    #
    #     'third_quarter_loiha_soni_reja',
    #     'third_quarter_loiha_soni_amalda',
    #     'third_quarter_umumiy_kiymati_reja',
    #     'third_quarter_umumiy_kiymati_amalda',
    #     'third_quarter_uz_mablag_reja',
    #     'third_quarter_uz_mablag_amalda',
    #     'third_quarter_bank_kredit_reja',
    #     'third_quarter_bank_kredit_amalda',
    #     'third_quarter_xorijiy_kredit_reja',
    #     'third_quarter_xorijiy_kredit_amalda',
    #     'third_quarter_xorijiy_invest_reja',
    #     'third_quarter_xorijiy_invest_amalda',
    #     'third_quarter_yangi_ish_reja',
    #     'third_quarter_yangi_ish_amalda',
    #     'third_quarter_ishlab_chiqarish_reja',
    #     'third_quarter_ishlab_chiqarish_amalda',
    #     'third_quarter_import_reja',
    #     'third_quarter_import_amalda',
    #     'third_quarter_export_reja',
    #     'third_quarter_export_amalda',
    #     'third_quarter_budget_reja',
    #     'third_quarter_budget_amalda',
    #
    #     'fourth_quarter_loiha_soni_reja',
    #     'fourth_quarter_loiha_soni_amalda',
    #     'fourth_quarter_umumiy_kiymati_reja',
    #     'fourth_quarter_umumiy_kiymati_amalda',
    #     'fourth_quarter_uz_mablag_reja',
    #     'fourth_quarter_uz_mablag_amalda',
    #     'fourth_quarter_bank_kredit_reja',
    #     'fourth_quarter_bank_kredit_amalda',
    #     'fourth_quarter_xorijiy_kredit_reja',
    #     'fourth_quarter_xorijiy_kredit_amalda',
    #     'fourth_quarter_xorijiy_invest_reja',
    #     'fourth_quarter_xorijiy_invest_amalda',
    #     'fourth_quarter_yangi_ish_reja',
    #     'fourth_quarter_yangi_ish_amalda',
    #     'fourth_quarter_ishlab_chiqarish_reja',
    #     'fourth_quarter_ishlab_chiqarish_amalda',
    #     'fourth_quarter_import_reja',
    #     'fourth_quarter_import_amalda',
    #     'fourth_quarter_export_reja',
    #     'fourth_quarter_export_amalda',
    #     'fourth_quarter_budget_reja',
    #     'fourth_quarter_budget_amalda',
    # ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_vault_tables(request, filter_slug, QuarterVault, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)

    wb.save(response)
    return response


from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment
import openpyxl


def export_excel_monthly(request, filter_slug):
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xlsx'

    wb = Workbook()
    ws = wb.active

    font_style = NamedStyle(name='font_style')
    font_style.font = Font(bold=True)

    cell_style = NamedStyle(name='cell_style')
    cell_style.font = Font(bold=True)
    cell_style.alignment = Alignment(vertical='center', horizontal='center')

    cell_title = NamedStyle(name='cell_title')
    cell_title.font = Font(bold=True, size=16)
    cell_title.alignment = Alignment(vertical='center', horizontal='left')

    row_num = 5  # с какой строки начинается наша таблица

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_vault_id(request):
        # ws.write_merge(0, 0, 0, 87, f'{request.user.district}, Cвод ойлар', cell_title)
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=264)
        ws.cell(row=1, column=1, value=f'{request.user.district}, Cвод ойлар').style = cell_title
        ws.merge_cells(start_row=2, end_row=4, start_column=1, end_column=2)
        ws.cell(row=2, column=1, value='Лойиҳа сони').style = cell_style
        ws.merge_cells(start_row=2, end_row=2, start_column=3, end_column=22)
        ws.cell(row=2, column=3, value='Январь').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=23, end_column=24)
        ws.cell(row=2, column=23, value='Лойиҳа сони').style = cell_style
        ws.merge_cells(start_row=2, end_row=2, start_column=25, end_column=44)
        ws.cell(row=2, column=25, value='Февраль').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=45, end_column=46)
        ws.cell(row=2, column=45, value='Лойиҳа сони').style = cell_style
        ws.merge_cells(start_row=2, end_row=2, start_column=47, end_column=66)
        ws.cell(row=2, column=47, value='Март').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=67, end_column=68)
        ws.cell(row=2, column=67, value='Лойиҳа сони').style = cell_style
        ws.merge_cells(start_row=2, end_row=2, start_column=69, end_column=88)
        ws.cell(row=2, column=69, value='Апрель').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=89, end_column=90)
        ws.cell(row=2, column=89, value='Лойиҳа сони').style = cell_style

        ws.merge_cells(start_row=2, end_row=2, start_column=91, end_column=110)
        ws.cell(row=2, column=91, value='Май').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=111, end_column=112)
        ws.cell(row=2, column=111, value='Лойиҳа сони').style = cell_style

        ws.merge_cells(start_row=2, end_row=2, start_column=113, end_column=132)
        ws.cell(row=2, column=113, value='Июль').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=133, end_column=134)
        ws.cell(row=2, column=133, value='Лойиҳа сони').style = cell_style

        ws.merge_cells(start_row=2, end_row=2, start_column=135, end_column=154)
        ws.cell(row=2, column=135, value='Июнь').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=155, end_column=156)
        ws.cell(row=2, column=155, value='Лойиҳа сони').style = cell_style

        ws.merge_cells(start_row=2, end_row=2, start_column=157, end_column=176)
        ws.cell(row=2, column=157, value='Август').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=177, end_column=178)
        ws.cell(row=2, column=177, value='Лойиҳа сони').style = cell_style

        ws.merge_cells(start_row=2, end_row=2, start_column=179, end_column=198)
        ws.cell(row=2, column=179, value='Сентябрь').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=199, end_column=200)
        ws.cell(row=2, column=199, value='Лойиҳа сони').style = cell_style

        ws.merge_cells(start_row=2, end_row=2, start_column=201, end_column=220)
        ws.cell(row=2, column=201, value='Октябрь').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=221, end_column=222)
        ws.cell(row=2, column=221, value='Лойиҳа сони').style = cell_style

        ws.merge_cells(start_row=2, end_row=2, start_column=223, end_column=242)
        ws.cell(row=2, column=223, value='Ноябрь').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=243, end_column=244)
        ws.cell(row=2, column=243, value='Лойиҳа сони').style = cell_style

        ws.merge_cells(start_row=2, end_row=2, start_column=245, end_column=264)
        ws.cell(row=2, column=245, value='Декабрь').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=3, end_column=4)
        ws.cell(row=3, column=3, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=5, end_column=12)
        ws.cell(row=3, column=5, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=13, end_column=14)
        ws.cell(row=3, column=13, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=15, end_column=22)
        ws.cell(row=3, column=15, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=25, end_column=26)
        ws.cell(row=3, column=25, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=27, end_column=34)
        ws.cell(row=3, column=27, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=35, end_column=36)
        ws.cell(row=3, column=35, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=37, end_column=44)
        ws.cell(row=3, column=37, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=47, end_column=48)
        ws.cell(row=3, column=47, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=49, end_column=56)
        ws.cell(row=3, column=49, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=57, end_column=58)
        ws.cell(row=3, column=57, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=59, end_column=66)
        ws.cell(row=3, column=59, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=69, end_column=70)
        ws.cell(row=3, column=69, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=71, end_column=78)
        ws.cell(row=3, column=71, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=79, end_column=80)
        ws.cell(row=3, column=79, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=81, end_column=88)
        ws.cell(row=3, column=81, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=91, end_column=92)
        ws.cell(row=3, column=91, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=93, end_column=100)
        ws.cell(row=3, column=93, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=101, end_column=102)
        ws.cell(row=3, column=101, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=103, end_column=110)
        ws.cell(row=3, column=103, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=113, end_column=114)
        ws.cell(row=3, column=113, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=115, end_column=122)
        ws.cell(row=3, column=115, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=123, end_column=124)
        ws.cell(row=3, column=123, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=125, end_column=132)
        ws.cell(row=3, column=125, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=135, end_column=136)
        ws.cell(row=3, column=135, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=137, end_column=144)
        ws.cell(row=3, column=137, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=145, end_column=146)
        ws.cell(row=3, column=145, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=147, end_column=154)
        ws.cell(row=3, column=147, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=157, end_column=158)
        ws.cell(row=3, column=157, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=159, end_column=166)
        ws.cell(row=3, column=159, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=167, end_column=168)
        ws.cell(row=3, column=167, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=169, end_column=176)
        ws.cell(row=3, column=169, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=179, end_column=180)
        ws.cell(row=3, column=179, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=181, end_column=188)
        ws.cell(row=3, column=181, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=189, end_column=190)
        ws.cell(row=3, column=189, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=191, end_column=198)
        ws.cell(row=3, column=191, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=201, end_column=202)
        ws.cell(row=3, column=201, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=203, end_column=210)
        ws.cell(row=3, column=203, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=211, end_column=212)
        ws.cell(row=3, column=211, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=213, end_column=220)
        ws.cell(row=3, column=213, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=223, end_column=224)
        ws.cell(row=3, column=223, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=225, end_column=232)
        ws.cell(row=3, column=225, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=233, end_column=234)
        ws.cell(row=3, column=233, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=235, end_column=242)
        ws.cell(row=3, column=235, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=245, end_column=246)
        ws.cell(row=3, column=245, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=247, end_column=254)
        ws.cell(row=3, column=247, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=255, end_column=256)
        ws.cell(row=3, column=255, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=257, end_column=264)
        ws.cell(row=3, column=257, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=5, end_column=6)
        ws.cell(row=4, column=5, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=7, end_column=8)
        ws.cell(row=4, column=7, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=9, end_column=10)
        ws.cell(row=4, column=9, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=11, end_column=12)
        ws.cell(row=4, column=11, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=15, end_column=16)
        ws.cell(row=4, column=15, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=17, end_column=18)
        ws.cell(row=4, column=17, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=19, end_column=20)
        ws.cell(row=4, column=19, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=21, end_column=22)
        ws.cell(row=4, column=21, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style
        # Февраль
        ws.merge_cells(start_row=4, end_row=4, start_column=27, end_column=28)
        ws.cell(row=4, column=27, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=29, end_column=30)
        ws.cell(row=4, column=29, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=31, end_column=32)
        ws.cell(row=4, column=31, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=33, end_column=34)
        ws.cell(row=4, column=33, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=37, end_column=38)
        ws.cell(row=4, column=37, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=39, end_column=40)
        ws.cell(row=4, column=39, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=41, end_column=42)
        ws.cell(row=4, column=41, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=43, end_column=44)
        ws.cell(row=4, column=43, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style

        # март
        ws.merge_cells(start_row=4, end_row=4, start_column=49, end_column=50)
        ws.cell(row=4, column=49, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=51, end_column=52)
        ws.cell(row=4, column=51, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=53, end_column=54)
        ws.cell(row=4, column=53, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=55, end_column=56)
        ws.cell(row=4, column=55, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=59, end_column=60)
        ws.cell(row=4, column=59, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=61, end_column=62)
        ws.cell(row=4, column=61, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=63, end_column=64)
        ws.cell(row=4, column=63, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=65, end_column=66)
        ws.cell(row=4, column=65, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style

        # апрель
        ws.merge_cells(start_row=4, end_row=4, start_column=71, end_column=72)
        ws.cell(row=4, column=71, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=73, end_column=74)
        ws.cell(row=4, column=73, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=75, end_column=76)
        ws.cell(row=4, column=75, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=77, end_column=78)
        ws.cell(row=4, column=77, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=81, end_column=82)
        ws.cell(row=4, column=81, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=83, end_column=84)
        ws.cell(row=4, column=83, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=85, end_column=86)
        ws.cell(row=4, column=85, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=87, end_column=88)
        ws.cell(row=4, column=87, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style

        # май
        ws.merge_cells(start_row=4, end_row=4, start_column=93, end_column=94)
        ws.cell(row=4, column=93, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=95, end_column=96)
        ws.cell(row=4, column=95, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=97, end_column=98)
        ws.cell(row=4, column=97, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=99, end_column=100)
        ws.cell(row=4, column=99, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=103, end_column=104)
        ws.cell(row=4, column=103, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=105, end_column=106)
        ws.cell(row=4, column=105, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=107, end_column=108)
        ws.cell(row=4, column=107, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=109, end_column=110)
        ws.cell(row=4, column=109, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style

        # июнь
        ws.merge_cells(start_row=4, end_row=4, start_column=115, end_column=116)
        ws.cell(row=4, column=115, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=117, end_column=118)
        ws.cell(row=4, column=117, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=119, end_column=120)
        ws.cell(row=4, column=119, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=121, end_column=122)
        ws.cell(row=4, column=121, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=125, end_column=126)
        ws.cell(row=4, column=125, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=127, end_column=128)
        ws.cell(row=4, column=127, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=129, end_column=130)
        ws.cell(row=4, column=129, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=131, end_column=132)
        ws.cell(row=4, column=131, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style


        # июль
        ws.merge_cells(start_row=4, end_row=4, start_column=137, end_column=138)
        ws.cell(row=4, column=137, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=139, end_column=140)
        ws.cell(row=4, column=139, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=141, end_column=142)
        ws.cell(row=4, column=141, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=143, end_column=144)
        ws.cell(row=4, column=143, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=147, end_column=148)
        ws.cell(row=4, column=147, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=149, end_column=150)
        ws.cell(row=4, column=149, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=151, end_column=152)
        ws.cell(row=4, column=151, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=153, end_column=154)
        ws.cell(row=4, column=153, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style

        # Август
        ws.merge_cells(start_row=4, end_row=4, start_column=159, end_column=160)
        ws.cell(row=4, column=159, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=161, end_column=162)
        ws.cell(row=4, column=161, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=163, end_column=164)
        ws.cell(row=4, column=163, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=165, end_column=166)
        ws.cell(row=4, column=165, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=169, end_column=170)
        ws.cell(row=4, column=169, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=171, end_column=172)
        ws.cell(row=4, column=171, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=173, end_column=174)
        ws.cell(row=4, column=173, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=175, end_column=176)
        ws.cell(row=4, column=175, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style


        # Сентябрь
        ws.merge_cells(start_row=4, end_row=4, start_column=181, end_column=182)
        ws.cell(row=4, column=181, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=183, end_column=184)
        ws.cell(row=4, column=183, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=185, end_column=186)
        ws.cell(row=4, column=185, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=187, end_column=188)
        ws.cell(row=4, column=187, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=191, end_column=192)
        ws.cell(row=4, column=191, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=193, end_column=194)
        ws.cell(row=4, column=193, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=195, end_column=196)
        ws.cell(row=4, column=195, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=197, end_column=198)
        ws.cell(row=4, column=197, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style

        # Октябрь
        ws.merge_cells(start_row=4, end_row=4, start_column=203, end_column=204)
        ws.cell(row=4, column=203, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=205, end_column=206)
        ws.cell(row=4, column=205, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=207, end_column=208)
        ws.cell(row=4, column=207, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=209, end_column=210)
        ws.cell(row=4, column=209, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=213, end_column=214)
        ws.cell(row=4, column=213, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=215, end_column=216)
        ws.cell(row=4, column=215, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=217, end_column=218)
        ws.cell(row=4, column=217, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=219, end_column=220)
        ws.cell(row=4, column=219, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style

        # Ноябрь
        ws.merge_cells(start_row=4, end_row=4, start_column=225, end_column=226)
        ws.cell(row=4, column=225, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=227, end_column=228)
        ws.cell(row=4, column=227, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=229, end_column=230)
        ws.cell(row=4, column=229, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=231, end_column=232)
        ws.cell(row=4, column=231, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=235, end_column=236)
        ws.cell(row=4, column=235, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=237, end_column=238)
        ws.cell(row=4, column=237, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=239, end_column=240)
        ws.cell(row=4, column=239, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=241, end_column=242)
        ws.cell(row=4, column=241, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style

        # Декабрь
        ws.merge_cells(start_row=4, end_row=4, start_column=247, end_column=248)
        ws.cell(row=4, column=247, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=249, end_column=250)
        ws.cell(row=4, column=249, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=251, end_column=252)
        ws.cell(row=4, column=251, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=253, end_column=254)
        ws.cell(row=4, column=253, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=257, end_column=258)
        ws.cell(row=4, column=257, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=259, end_column=260)
        ws.cell(row=4, column=259, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=261, end_column=262)
        ws.cell(row=4, column=261, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=263, end_column=264)
        ws.cell(row=4, column=263, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style
        # ws.merge_cells(start_row=2, end_row=2, start_column=21, end_column=35)

    else:
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=265)
        ws.cell(row=1, column=1, value='Cвод ойлар').style = cell_title
        ws.merge_cells(start_row=2, end_row=4, start_column=2, end_column=3)
        ws.cell(row=2, column=2, value='Лойиҳа сони').style = cell_style
        ws.merge_cells(start_row=2, end_row=2, start_column=4, end_column=23)
        ws.cell(row=2, column=4, value='Январь').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=24, end_column=25)
        ws.cell(row=2, column=24, value='Лойиҳа сони').style = cell_style
        ws.merge_cells(start_row=2, end_row=2, start_column=26, end_column=45)
        ws.cell(row=2, column=26, value='Февраль').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=46, end_column=47)
        ws.cell(row=2, column=46, value='Лойиҳа сони').style = cell_style
        ws.merge_cells(start_row=2, end_row=2, start_column=48, end_column=67)
        ws.cell(row=2, column=48, value='Март').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=68, end_column=69)
        ws.cell(row=2, column=68, value='Лойиҳа сони').style = cell_style
        ws.merge_cells(start_row=2, end_row=2, start_column=70, end_column=89)
        ws.cell(row=2, column=70, value='Апрель').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=90, end_column=91)
        ws.cell(row=2, column=90, value='Лойиҳа сони').style = cell_style

        ws.merge_cells(start_row=2, end_row=2, start_column=92, end_column=111)
        ws.cell(row=2, column=92, value='Май').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=112, end_column=113)
        ws.cell(row=2, column=112, value='Лойиҳа сони').style = cell_style

        ws.merge_cells(start_row=2, end_row=2, start_column=114, end_column=133)
        ws.cell(row=2, column=114, value='Июль').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=134, end_column=135)
        ws.cell(row=2, column=134, value='Лойиҳа сони').style = cell_style

        ws.merge_cells(start_row=2, end_row=2, start_column=136, end_column=155)
        ws.cell(row=2, column=136, value='Июнь').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=156, end_column=157)
        ws.cell(row=2, column=156, value='Лойиҳа сони').style = cell_style

        ws.merge_cells(start_row=2, end_row=2, start_column=158, end_column=177)
        ws.cell(row=2, column=158, value='Август').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=178, end_column=179)
        ws.cell(row=2, column=178, value='Лойиҳа сони').style = cell_style

        ws.merge_cells(start_row=2, end_row=2, start_column=180, end_column=199)
        ws.cell(row=2, column=180, value='Сентябрь').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=200, end_column=201)
        ws.cell(row=2, column=200, value='Лойиҳа сони').style = cell_style

        ws.merge_cells(start_row=2, end_row=2, start_column=202, end_column=221)
        ws.cell(row=2, column=202, value='Октябрь').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=222, end_column=223)
        ws.cell(row=2, column=222, value='Лойиҳа сони').style = cell_style

        ws.merge_cells(start_row=2, end_row=2, start_column=224, end_column=243)
        ws.cell(row=2, column=224, value='Ноябрь').style = cell_style
        ws.merge_cells(start_row=2, end_row=4, start_column=244, end_column=245)
        ws.cell(row=2, column=244, value='Лойиҳа сони').style = cell_style

        ws.merge_cells(start_row=2, end_row=2, start_column=246, end_column=265)
        ws.cell(row=2, column=246, value='Декабрь').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=4, end_column=5)
        ws.cell(row=3, column=4, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=6, end_column=13)
        ws.cell(row=3, column=6, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=14, end_column=15)
        ws.cell(row=3, column=14, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=16, end_column=23)
        ws.cell(row=3, column=16, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=26, end_column=27)
        ws.cell(row=3, column=26, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=28, end_column=35)
        ws.cell(row=3, column=28, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=36, end_column=37)
        ws.cell(row=3, column=36, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=38, end_column=45)
        ws.cell(row=3, column=38, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=48, end_column=49)
        ws.cell(row=3, column=48, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=50, end_column=57)
        ws.cell(row=3, column=50, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=58, end_column=59)
        ws.cell(row=3, column=58, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=60, end_column=67)
        ws.cell(row=3, column=60, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=70, end_column=71)
        ws.cell(row=3, column=70, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=72, end_column=79)
        ws.cell(row=3, column=72, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=80, end_column=81)
        ws.cell(row=3, column=80, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=82, end_column=89)
        ws.cell(row=3, column=82, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=92, end_column=93)
        ws.cell(row=3, column=92, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=94, end_column=101)
        ws.cell(row=3, column=94, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=102, end_column=103)
        ws.cell(row=3, column=102, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=104, end_column=111)
        ws.cell(row=3, column=104, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=114, end_column=115)
        ws.cell(row=3, column=114, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=116, end_column=123)
        ws.cell(row=3, column=116, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=124, end_column=125)
        ws.cell(row=3, column=124, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=126, end_column=133)
        ws.cell(row=3, column=126, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=136, end_column=137)
        ws.cell(row=3, column=136, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=138, end_column=145)
        ws.cell(row=3, column=138, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=146, end_column=147)
        ws.cell(row=3, column=146, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=148, end_column=155)
        ws.cell(row=3, column=148, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=158, end_column=159)
        ws.cell(row=3, column=158, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=160, end_column=167)
        ws.cell(row=3, column=160, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=168, end_column=169)
        ws.cell(row=3, column=168, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=170, end_column=177)
        ws.cell(row=3, column=170, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=180, end_column=181)
        ws.cell(row=3, column=180, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=182, end_column=189)
        ws.cell(row=3, column=182, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=190, end_column=191)
        ws.cell(row=3, column=190, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=192, end_column=199)
        ws.cell(row=3, column=192, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=202, end_column=203)
        ws.cell(row=3, column=202, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=204, end_column=211)
        ws.cell(row=3, column=204, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=212, end_column=213)
        ws.cell(row=3, column=212, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=214, end_column=221)
        ws.cell(row=3, column=214, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=224, end_column=225)
        ws.cell(row=3, column=224, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=226, end_column=233)
        ws.cell(row=3, column=226, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=234, end_column=235)
        ws.cell(row=3, column=234, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=236, end_column=243)
        ws.cell(row=3, column=236, value='Иқтисодий самарадорлик').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=246, end_column=247)
        ws.cell(row=3, column=246, value='Умумий қиймати млн.сўм').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=248, end_column=255)
        ws.cell(row=3, column=248, value='шундан').style = cell_style

        ws.merge_cells(start_row=3, end_row=4, start_column=256, end_column=257)
        ws.cell(row=3, column=256, value='Янги иш ўринлари сони').style = cell_style

        ws.merge_cells(start_row=3, end_row=3, start_column=258, end_column=265)
        ws.cell(row=3, column=258, value='Иқтисодий самарадорлик').style = cell_style


        ws.merge_cells(start_row=4, end_row=4, start_column=6, end_column=7)
        ws.cell(row=4, column=6, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=8, end_column=9)
        ws.cell(row=4, column=8, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=10, end_column=11)
        ws.cell(row=4, column=10, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=12, end_column=13)
        ws.cell(row=4, column=12, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=16, end_column=17)
        ws.cell(row=4, column=16, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=18, end_column=19)
        ws.cell(row=4, column=18, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=20, end_column=21)
        ws.cell(row=4, column=20, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=22, end_column=23)
        ws.cell(row=4, column=22, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style
        # Февраль
        ws.merge_cells(start_row=4, end_row=4, start_column=28, end_column=29)
        ws.cell(row=4, column=28, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=30, end_column=31)
        ws.cell(row=4, column=30, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=32, end_column=33)
        ws.cell(row=4, column=32, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=34, end_column=35)
        ws.cell(row=4, column=34, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=38, end_column=39)
        ws.cell(row=4, column=38, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=40, end_column=41)
        ws.cell(row=4, column=40, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=42, end_column=43)
        ws.cell(row=4, column=42, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=44, end_column=45)
        ws.cell(row=4, column=44, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style

        # март
        ws.merge_cells(start_row=4, end_row=4, start_column=50, end_column=51)
        ws.cell(row=4, column=50, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=52, end_column=53)
        ws.cell(row=4, column=52, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=54, end_column=55)
        ws.cell(row=4, column=54, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=56, end_column=57)
        ws.cell(row=4, column=56, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=60, end_column=61)
        ws.cell(row=4, column=60, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=62, end_column=63)
        ws.cell(row=4, column=62, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=64, end_column=65)
        ws.cell(row=4, column=64, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=66, end_column=67)
        ws.cell(row=4, column=66, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style

        # апрель
        ws.merge_cells(start_row=4, end_row=4, start_column=72, end_column=73)
        ws.cell(row=4, column=72, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=74, end_column=75)
        ws.cell(row=4, column=74, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=76, end_column=77)
        ws.cell(row=4, column=76, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=78, end_column=79)
        ws.cell(row=4, column=78, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=82, end_column=83)
        ws.cell(row=4, column=82, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=84, end_column=85)
        ws.cell(row=4, column=84, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=86, end_column=87)
        ws.cell(row=4, column=86, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=88, end_column=89)
        ws.cell(row=4, column=88, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style

        # май
        ws.merge_cells(start_row=4, end_row=4, start_column=94, end_column=95)
        ws.cell(row=4, column=94, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=96, end_column=97)
        ws.cell(row=4, column=96, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=98, end_column=99)
        ws.cell(row=4, column=98, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=100, end_column=101)
        ws.cell(row=4, column=100, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=104, end_column=105)
        ws.cell(row=4, column=104, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=106, end_column=107)
        ws.cell(row=4, column=106, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=108, end_column=109)
        ws.cell(row=4, column=108, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=110, end_column=111)
        ws.cell(row=4, column=110, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style

        # июнь
        ws.merge_cells(start_row=4, end_row=4, start_column=116, end_column=117)
        ws.cell(row=4, column=116, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=118, end_column=119)
        ws.cell(row=4, column=118, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=120, end_column=121)
        ws.cell(row=4, column=120, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=122, end_column=123)
        ws.cell(row=4, column=122, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=126, end_column=127)
        ws.cell(row=4, column=126, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=128, end_column=129)
        ws.cell(row=4, column=128, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=130, end_column=131)
        ws.cell(row=4, column=130, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=132, end_column=133)
        ws.cell(row=4, column=132, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style

        # июль
        ws.merge_cells(start_row=4, end_row=4, start_column=138, end_column=139)
        ws.cell(row=4, column=138, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=140, end_column=141)
        ws.cell(row=4, column=140, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=142, end_column=143)
        ws.cell(row=4, column=142, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=144, end_column=145)
        ws.cell(row=4, column=144, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=148, end_column=149)
        ws.cell(row=4, column=148, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=150, end_column=151)
        ws.cell(row=4, column=150, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=152, end_column=153)
        ws.cell(row=4, column=152, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=154, end_column=155)
        ws.cell(row=4, column=154, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style

        # Август
        ws.merge_cells(start_row=4, end_row=4, start_column=160, end_column=161)
        ws.cell(row=4, column=160, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=162, end_column=163)
        ws.cell(row=4, column=162, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=164, end_column=165)
        ws.cell(row=4, column=164, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=166, end_column=167)
        ws.cell(row=4, column=166, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=170, end_column=171)
        ws.cell(row=4, column=170, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=172, end_column=173)
        ws.cell(row=4, column=172, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=174, end_column=175)
        ws.cell(row=4, column=174, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=176, end_column=177)
        ws.cell(row=4, column=176, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style

        # Сентябрь
        ws.merge_cells(start_row=4, end_row=4, start_column=182, end_column=183)
        ws.cell(row=4, column=182, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=184, end_column=185)
        ws.cell(row=4, column=184, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=186, end_column=187)
        ws.cell(row=4, column=186, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=188, end_column=189)
        ws.cell(row=4, column=188, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=192, end_column=193)
        ws.cell(row=4, column=192, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=194, end_column=195)
        ws.cell(row=4, column=194, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=196, end_column=197)
        ws.cell(row=4, column=196, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=198, end_column=199)
        ws.cell(row=4, column=198, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style

        # Октябрь
        ws.merge_cells(start_row=4, end_row=4, start_column=204, end_column=205)
        ws.cell(row=4, column=204, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=206, end_column=207)
        ws.cell(row=4, column=206, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=208, end_column=209)
        ws.cell(row=4, column=208, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=210, end_column=211)
        ws.cell(row=4, column=210, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=214, end_column=215)
        ws.cell(row=4, column=214, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=216, end_column=217)
        ws.cell(row=4, column=216, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=218, end_column=219)
        ws.cell(row=4, column=218, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=220, end_column=221)
        ws.cell(row=4, column=220, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style

        # Ноябрь
        ws.merge_cells(start_row=4, end_row=4, start_column=226, end_column=227)
        ws.cell(row=4, column=226, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=228, end_column=229)
        ws.cell(row=4, column=228, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=230, end_column=231)
        ws.cell(row=4, column=230, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=232, end_column=233)
        ws.cell(row=4, column=232, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=236, end_column=237)
        ws.cell(row=4, column=236, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=238, end_column=239)
        ws.cell(row=4, column=238, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=240, end_column=241)
        ws.cell(row=4, column=240, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=242, end_column=243)
        ws.cell(row=4, column=242, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style

        # Декабрь
        ws.merge_cells(start_row=4, end_row=4, start_column=248, end_column=249)
        ws.cell(row=4, column=248, value='ўз маблағлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=250, end_column=251)
        ws.cell(row=4, column=250, value='банк кредитлари млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=252, end_column=253)
        ws.cell(row=4, column=252, value='хорижий кредитлар минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=254, end_column=255)
        ws.cell(row=4, column=254, value='хорижий инвестициялар минг.долл').style = cell_style

        ws.merge_cells(start_row=4, end_row=4, start_column=258, end_column=259)
        ws.cell(row=4, column=258, value='Ишлаб чиқариш қуввати млн.сўм').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=260, end_column=261)
        ws.cell(row=4, column=260, value='Импорт урнини босиш минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=262, end_column=263)
        ws.cell(row=4, column=262, value='Экспорт ҳажми минг.долл').style = cell_style
        ws.merge_cells(start_row=4, end_row=4, start_column=264, end_column=265)
        ws.cell(row=4, column=264, value='Бюджетга қўшимча тушум млн.сўм').style = cell_style

    # "Район",
    columns_list = [
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
        'Режa',
        'Амалда',
    ]

    if not get_group_vault_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        column_letter = openpyxl.utils.get_column_letter(col_num + 1)
        ws.column_dimensions[column_letter].width = 25
        ws.cell(row=row_num, column=col_num + 1, value=columns[col_num])
        ws.cell(row=row_num, column=col_num + 1).style = cell_style

    form = TableMonthForm()
    all_fields = form.fields.keys()
    fields = [
        *all_fields
    ]


    # fields = [
    #     'january_loiha_soni_reja',
    #     'january_loiha_soni_amalda',
    #     'january_umumiy_kiymati_reja',
    #     'january_umumiy_kiymati_amalda',
    #     'january_uz_mablag_reja',
    #     'january_uz_mablag_amalda',
    #     'january_bank_kredit_reja',
    #     'january_bank_kredit_amalda',
    #     'january_xorijiy_kredit_reja',
    #     'january_xorijiy_kredit_amalda',
    #     'january_xorijiy_invest_reja',
    #     'january_xorijiy_invest_amalda',
    #     'january_yangi_ish_reja',
    #     'january_yangi_ish_amalda',
    #     'january_ishlab_chiqarish_reja',
    #     'january_ishlab_chiqarish_amalda',
    #     'january_import_reja',
    #     'january_import_amalda',
    #     'january_export_reja',
    #     'january_export_amalda',
    #     'january_budget_reja',
    #     'january_budget_amalda',
    #
    #     'february_loiha_soni_reja',
    #     'february_loiha_soni_amalda',
    #     'february_umumiy_kiymati_reja',
    #     'february_umumiy_kiymati_amalda',
    #     'february_uz_mablag_reja',
    #     'february_uz_mablag_amalda',
    #     'february_bank_kredit_reja',
    #     'february_bank_kredit_amalda',
    #     'february_xorijiy_kredit_reja',
    #     'february_xorijiy_kredit_amalda',
    #     'february_xorijiy_invest_reja',
    #     'february_xorijiy_invest_amalda',
    #     'february_yangi_ish_reja',
    #     'february_yangi_ish_amalda',
    #     'february_ishlab_chiqarish_reja',
    #     'february_ishlab_chiqarish_amalda',
    #     'february_import_reja',
    #     'february_import_amalda',
    #     'february_export_reja',
    #     'february_export_amalda',
    #     'february_budget_reja',
    #     'february_budget_amalda',
    #
    #     'march_loiha_soni_reja',
    #     'march_loiha_soni_amalda',
    #     'march_umumiy_kiymati_reja',
    #     'march_umumiy_kiymati_amalda',
    #     'march_uz_mablag_reja',
    #     'march_uz_mablag_amalda',
    #     'march_bank_kredit_reja',
    #     'march_bank_kredit_amalda',
    #     'march_xorijiy_kredit_reja',
    #     'march_xorijiy_kredit_amalda',
    #     'march_xorijiy_invest_reja',
    #     'march_xorijiy_invest_amalda',
    #     'march_yangi_ish_reja',
    #     'march_yangi_ish_amalda',
    #     'march_ishlab_chiqarish_reja',
    #     'march_ishlab_chiqarish_amalda',
    #     'march_import_reja',
    #     'march_import_amalda',
    #     'march_export_reja',
    #     'march_export_amalda',
    #     'march_budget_reja',
    #     'march_budget_amalda',
    #
    #     'april_loiha_soni_reja',
    #     'april_loiha_soni_amalda',
    #     'april_umumiy_kiymati_reja',
    #     'april_umumiy_kiymati_amalda',
    #     'april_uz_mablag_reja',
    #     'april_uz_mablag_amalda',
    #     'april_bank_kredit_reja',
    #     'april_bank_kredit_amalda',
    #     'april_xorijiy_kredit_reja',
    #     'april_xorijiy_kredit_amalda',
    #     'april_xorijiy_invest_reja',
    #     'april_xorijiy_invest_amalda',
    #     'april_yangi_ish_reja',
    #     'april_yangi_ish_amalda',
    #     'april_ishlab_chiqarish_reja',
    #     'april_ishlab_chiqarish_amalda',
    #     'april_import_reja',
    #     'april_import_amalda',
    #     'april_export_reja',
    #     'april_export_amalda',
    #     'april_budget_reja',
    #     'april_budget_amalda',
    #
    #     'may_loiha_soni_reja',
    #     'may_loiha_soni_amalda',
    #     'may_umumiy_kiymati_reja',
    #     'may_umumiy_kiymati_amalda',
    #     'may_uz_mablag_reja',
    #     'may_uz_mablag_amalda',
    #     'may_bank_kredit_reja',
    #     'may_bank_kredit_amalda',
    #     'may_xorijiy_kredit_reja',
    #     'may_xorijiy_kredit_amalda',
    #     'may_xorijiy_invest_reja',
    #     'may_xorijiy_invest_amalda',
    #     'may_yangi_ish_reja',
    #     'may_yangi_ish_amalda',
    #     'may_ishlab_chiqarish_reja',
    #     'may_ishlab_chiqarish_amalda',
    #     'may_import_reja',
    #     'may_import_amalda',
    #     'may_export_reja',
    #     'may_export_amalda',
    #     'may_budget_reja',
    #     'may_budget_amalda',
    #
    #     'june_loiha_soni_reja',
    #     'june_loiha_soni_amalda',
    #     'june_umumiy_kiymati_reja',
    #     'june_umumiy_kiymati_amalda',
    #     'june_uz_mablag_reja',
    #     'june_uz_mablag_amalda',
    #     'june_bank_kredit_reja',
    #     'june_bank_kredit_amalda',
    #     'june_xorijiy_kredit_reja',
    #     'june_xorijiy_kredit_amalda',
    #     'june_xorijiy_invest_reja',
    #     'june_xorijiy_invest_amalda',
    #     'june_yangi_ish_reja',
    #     'june_yangi_ish_amalda',
    #     'june_ishlab_chiqarish_reja',
    #     'june_ishlab_chiqarish_amalda',
    #     'june_import_reja',
    #     'june_import_amalda',
    #     'june_export_reja',
    #     'june_export_amalda',
    #     'june_budget_reja',
    #     'june_budget_amalda',
    #
    #     'july_loiha_soni_reja',
    #     'july_loiha_soni_amalda',
    #     'july_umumiy_kiymati_reja',
    #     'july_umumiy_kiymati_amalda',
    #     'july_uz_mablag_reja',
    #     'july_uz_mablag_amalda',
    #     'july_bank_kredit_reja',
    #     'july_bank_kredit_amalda',
    #     'july_xorijiy_kredit_reja',
    #     'july_xorijiy_kredit_amalda',
    #     'july_xorijiy_invest_reja',
    #     'july_xorijiy_invest_amalda',
    #     'july_yangi_ish_reja',
    #     'july_yangi_ish_amalda',
    #     'july_ishlab_chiqarish_reja',
    #     'july_ishlab_chiqarish_amalda',
    #     'july_import_reja',
    #     'july_import_amalda',
    #     'july_export_reja',
    #     'july_export_amalda',
    #     'july_budget_reja',
    #     'july_budget_amalda',
    #
    #     'august_loiha_soni_reja',
    #     'august_loiha_soni_amalda',
    #     'august_umumiy_kiymati_reja',
    #     'august_umumiy_kiymati_amalda',
    #     'august_uz_mablag_reja',
    #     'august_uz_mablag_amalda',
    #     'august_bank_kredit_reja',
    #     'august_bank_kredit_amalda',
    #     'august_xorijiy_kredit_reja',
    #     'august_xorijiy_kredit_amalda',
    #     'august_xorijiy_invest_reja',
    #     'august_xorijiy_invest_amalda',
    #     'august_yangi_ish_reja',
    #     'august_yangi_ish_amalda',
    #     'august_ishlab_chiqarish_reja',
    #     'august_ishlab_chiqarish_amalda',
    #     'august_import_reja',
    #     'august_import_amalda',
    #     'august_export_reja',
    #     'august_export_amalda',
    #     'august_budget_reja',
    #     'august_budget_amalda',
    #
    #     'september_loiha_soni_reja',
    #     'september_loiha_soni_amalda',
    #     'september_umumiy_kiymati_reja',
    #     'september_umumiy_kiymati_amalda',
    #     'september_uz_mablag_reja',
    #     'september_uz_mablag_amalda',
    #     'september_bank_kredit_reja',
    #     'september_bank_kredit_amalda',
    #     'september_xorijiy_kredit_reja',
    #     'september_xorijiy_kredit_amalda',
    #     'september_xorijiy_invest_reja',
    #     'september_xorijiy_invest_amalda',
    #     'september_yangi_ish_reja',
    #     'september_yangi_ish_amalda',
    #     'september_ishlab_chiqarish_reja',
    #     'september_ishlab_chiqarish_amalda',
    #     'september_import_reja',
    #     'september_import_amalda',
    #     'september_export_reja',
    #     'september_export_amalda',
    #     'september_budget_reja',
    #     'september_budget_amalda',
    #
    #     'october_loiha_soni_reja',
    #     'october_loiha_soni_amalda',
    #     'october_umumiy_kiymati_reja',
    #     'october_umumiy_kiymati_amalda',
    #     'october_uz_mablag_reja',
    #     'october_uz_mablag_amalda',
    #     'october_bank_kredit_reja',
    #     'october_bank_kredit_amalda',
    #     'october_xorijiy_kredit_reja',
    #     'october_xorijiy_kredit_amalda',
    #     'october_xorijiy_invest_reja',
    #     'october_xorijiy_invest_amalda',
    #     'october_yangi_ish_reja',
    #     'october_yangi_ish_amalda',
    #     'october_ishlab_chiqarish_reja',
    #     'october_ishlab_chiqarish_amalda',
    #     'october_import_reja',
    #     'october_import_amalda',
    #     'october_export_reja',
    #     'october_export_amalda',
    #     'october_budget_reja',
    #     'october_budget_amalda',
    #
    #     'november_loiha_soni_reja',
    #     'november_loiha_soni_amalda',
    #     'november_umumiy_kiymati_reja',
    #     'november_umumiy_kiymati_amalda',
    #     'november_uz_mablag_reja',
    #     'november_uz_mablag_amalda',
    #     'november_bank_kredit_reja',
    #     'november_bank_kredit_amalda',
    #     'november_xorijiy_kredit_reja',
    #     'november_xorijiy_kredit_amalda',
    #     'november_xorijiy_invest_reja',
    #     'november_xorijiy_invest_amalda',
    #     'november_yangi_ish_reja',
    #     'november_yangi_ish_amalda',
    #     'november_ishlab_chiqarish_reja',
    #     'november_ishlab_chiqarish_amalda',
    #     'november_import_reja',
    #     'november_import_amalda',
    #     'november_export_reja',
    #     'november_export_amalda',
    #     'november_budget_reja',
    #     'november_budget_amalda',
    #
    #     'december_loiha_soni_reja',
    #     'december_loiha_soni_amalda',
    #     'december_umumiy_kiymati_reja',
    #     'december_umumiy_kiymati_amalda',
    #     'december_uz_mablag_reja',
    #     'december_uz_mablag_amalda',
    #     'december_bank_kredit_reja',
    #     'december_bank_kredit_amalda',
    #     'december_xorijiy_kredit_reja',
    #     'december_xorijiy_kredit_amalda',
    #     'december_xorijiy_invest_reja',
    #     'december_xorijiy_invest_amalda',
    #     'december_yangi_ish_reja',
    #     'december_yangi_ish_amalda',
    #     'december_ishlab_chiqarish_reja',
    #     'december_ishlab_chiqarish_amalda',
    #     'december_import_reja',
    #     'december_import_amalda',
    #     'december_export_reja',
    #     'december_export_amalda',
    #     'december_budget_reja',
    #     'december_budget_amalda',
    # ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_vault_tables(request, filter_slug, MonthVault, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.cell(row=row_num, column=col_num + 1, value=str(row[col_num]))

    wb.save(response)
    return response

# 4 отдел постмонитлоринга

def export_excel_addressed(request, filter_slug):
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xlsx'

    wb = Workbook()
    ws = wb.active

    font_style = NamedStyle(name='font_style')
    font_style.font = Font(bold=True)

    cell_style = NamedStyle(name='cell_style')
    cell_style.font = Font(bold=True)
    cell_style.alignment = Alignment(vertical='center', horizontal='left')

    cell_title = NamedStyle(name='cell_title')
    cell_title.font = Font(bold=True, size=16)
    cell_title.alignment = Alignment(vertical='center', horizontal='left')

    row_num = 3  # с какой строки начинается наша таблица

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_post_monitoring_id(request):
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=36)
        ws.cell(row=1, column=1, value=f'{request.user.district}, Манзилли').style = cell_title
        ws.merge_cells(start_row=2, end_row=2, start_column=14, end_column=15)
        ws.cell(row=2, column=14, value='Хорижий ҳамкор').style = cell_style
        ws.merge_cells(start_row=2, end_row=2, start_column=18, end_column=19)
        ws.cell(row=2, column=18, value='Йиллик қуввати (ТИА/бизнес режа бўйича)').style = cell_style
        ws.merge_cells(start_row=2, end_row=2, start_column=21, end_column=22)
        ws.cell(row=2, column=21, value='2023 йил январь ойида режалаштирилган ишлаб чиқариш ҳажмлар').style = cell_style
        ws.merge_cells(start_row=2, end_row=2, start_column=23, end_column=24)
        ws.cell(row=2, column=23, value='2023 йил январь ойида ишлаб чиқарилган маҳсулот ҳажми').style = cell_style
        ws.merge_cells(start_row=2, end_row=2, start_column=25, end_column=26)
        ws.cell(row=2, column=25, value='2023 йил январь ойида экспорт ҳажми').style = cell_style
        ws.merge_cells(start_row=2, end_row=2, start_column=27, end_column=28)
        ws.cell(row=2, column=27, value='2023 йил январь ойида бюджет тушумлари ҳажми').style = cell_style
        ws.merge_cells(start_row=2, end_row=2, start_column=29, end_column=30)
        ws.cell(row=2, column=29, value='Иш ўринлари бандлиги').style = cell_style

        # ws.merge_cells(start_row=2, end_row=2, start_column=21, end_column=35)

    else:
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=37)
        ws.cell(row=1, column=1, value='Манзилли').style = cell_title

        ws.merge_cells(start_row=2, end_row=2, start_column=15, end_column=16)
        ws.cell(row=2, column=15, value='Хорижий ҳамкор').style = cell_style
        ws.merge_cells(start_row=2, end_row=2, start_column=19, end_column=20)
        ws.cell(row=2, column=19, value='Йиллик қуввати (ТИА/бизнес режа бўйича)').style = cell_style
        ws.merge_cells(start_row=2, end_row=2, start_column=22, end_column=23)
        ws.cell(row=2, column=22, value='2023 йил январь ойида режалаштирилган ишлаб чиқариш ҳажмлар').style = cell_style
        ws.merge_cells(start_row=2, end_row=2, start_column=24, end_column=25)
        ws.cell(row=2, column=24, value='2023 йил январь ойида ишлаб чиқарилган маҳсулот ҳажми').style = cell_style
        ws.merge_cells(start_row=2, end_row=2, start_column=26, end_column=27)
        ws.cell(row=2, column=26, value='2023 йил январь ойида экспорт ҳажми').style = cell_style
        ws.merge_cells(start_row=2, end_row=2, start_column=28, end_column=29)
        ws.cell(row=2, column=28, value='2023 йил январь ойида бюджет тушумлари ҳажми').style = cell_style
        ws.merge_cells(start_row=2, end_row=2, start_column=30, end_column=31)
        ws.cell(row=2, column=30, value='Иш ўринлари бандлиги').style = cell_style


    # "Район",
    columns_list = [
        'Туман номи',
        'Масъул ташкилот',
        'Корхона номи',
        'ТЕГМА',
        'Лойиҳа қиймати млн.долл.',
        'лойиҳа ташкил этилиши (янгидан, модернизация, кенгайтириш)',
        'Йўналишлар номи',
        'Тармоқлар номи',
        'ИНН рақами',
        'Юридик (жойлашган) манзил',
        'Корхона раҳбари',
        'Телефон рақами',
        'Кафолатлилиги (Давлат, Тўғридан-тўғри хорижий инвестиция, хорижий кредит, ўз маблағи)',
        'Давлат',
        'Компания номи',
        'Йил',
        'Лойиҳа ишга тушган даври (кун, ой, йил)',
        '(натурада)',
        '(млн. сўмда)',
        'Яратилган иш ўрни сони',
        'натурада',
        'млн.сўм',
        'натурада',
        'млн.сўм',
        'Режа (минг доллар)',
        'Амалда (минг доллар)',
        'Режа (млн сўм)',
        'Амалда (млн сўм)',
        'Ишчилар сони',
        'Вакансия сони',
        'Хизмат кўрсатувчи банк',
        'Корхона холати (Инфратузилма объекти, Мавсумий, Вақтинча ишламаяпти, Тугатилган, Тўлиқ қувват (80-100%), Ўрта қувват (40-79%), Паст қувват (0-39,9%))',
        'Кам қувватда ишлаш сабаби (бу устунда фақат паст қувват бўлса ёзилади)',
        'Муаммо мавжудлиги (бўлмаса бўш қолсин, мавжуд эмас)',
        'Муаммони бартараф этиш бўйича таклиф',
        'Муаммони ҳал этишга масъул ташкилот',
    ]

    if not get_group_post_monitoring_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        column_letter = openpyxl.utils.get_column_letter(col_num + 1)
        ws.column_dimensions[column_letter].width = 25
        ws.cell(row=row_num, column=col_num + 1, value=columns[col_num])
        ws.cell(row=row_num, column=col_num + 1).style = cell_style

    form = TableAddressedForm()
    all_fields = form.fields.keys()
    fields = [
        *all_fields
    ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_postmonitoring_tables(request, filter_slug, Addressed, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.cell(row=row_num, column=col_num + 1, value=str(row[col_num]))

    wb.save(response)
    return response


def export_excel_subtotals(request, filter_slug):
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xlsx'

    wb = Workbook()
    ws = wb.active

    font_style = NamedStyle(name='font_style')
    font_style.font = Font(bold=True)

    cell_style = NamedStyle(name='cell_style')
    cell_style.font = Font(bold=True)
    cell_style.alignment = Alignment(vertical='center', horizontal='left')

    cell_title = NamedStyle(name='cell_title')
    cell_title.font = Font(bold=True, size=16)
    cell_title.alignment = Alignment(vertical='center', horizontal='left')

    row_num = 3  # с какой строки начинается наша таблица

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_post_monitoring_id(request):
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=8)
        ws.cell(row=1, column=1, value=f'{request.user.district}, Промежуточный итоги').style = cell_title
        ws.merge_cells(start_row=2, end_row=2, start_column=2, end_column=8)
        ws.cell(row=2, column=2, value='Январь ойи').style = cell_style
    else:
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=9)
        ws.cell(row=1, column=1, value='Промежуточный итоги').style = cell_title
        ws.merge_cells(start_row=2, end_row=2, start_column=3, end_column=9)
        ws.cell(row=2, column=3, value='Хорижий ҳамкор').style = cell_style


    # "Район",
    columns_list = [
        'Умумий корхоналар сони',
        'Тўлиқ қувват (80-100%)',
        'Ўрта қувват (51-79%)',
        'Паст қувват (0-50%)',
        'Вақтинча ишламаяпти',
        'Тугатилган',
        'Мавсумий',
        'Бошқа',
    ]

    if not get_group_post_monitoring_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        column_letter = openpyxl.utils.get_column_letter(col_num + 1)
        ws.column_dimensions[column_letter].width = 25
        ws.cell(row=row_num, column=col_num + 1, value=columns[col_num])
        ws.cell(row=row_num, column=col_num + 1).style = cell_style

    form = TableSubtotalsForm()
    all_fields = form.fields.keys()
    fields = [
        *all_fields
    ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_postmonitoring_tables(request, filter_slug, Subtotals, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.cell(row=row_num, column=col_num + 1, value=str(row[col_num]))

    wb.save(response)
    return response


def export_excel_(request, filter_slug):
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename=table{str(datetime.now())}.xlsx'

    wb = Workbook()
    ws = wb.active

    font_style = NamedStyle(name='font_style')
    font_style.font = Font(bold=True)

    cell_style = NamedStyle(name='cell_style')
    cell_style.font = Font(bold=True)
    cell_style.alignment = Alignment(vertical='center', horizontal='left')

    cell_title = NamedStyle(name='cell_title')
    cell_title.font = Font(bold=True, size=16)
    cell_title.alignment = Alignment(vertical='center', horizontal='left')

    row_num = 3  # с какой строки начинается наша таблица

    # ws.title = 'Илова-4.1' Поменять на один если добавлять столбик в начале
    if not get_group_post_monitoring_id(request):
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=8)
        ws.cell(row=1, column=1, value=f'{request.user.district}, Промежуточный итоги').style = cell_title
        ws.merge_cells(start_row=2, end_row=2, start_column=2, end_column=8)
        ws.cell(row=2, column=2, value='Январь ойи').style = cell_style
    else:
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=9)
        ws.cell(row=1, column=1, value='Промежуточный итоги').style = cell_title
        ws.merge_cells(start_row=2, end_row=2, start_column=3, end_column=9)
        ws.cell(row=2, column=3, value='Хорижий ҳамкор').style = cell_style


    # "Район",
    columns_list = [
        'Умумий корхоналар сони',
        'Тўлиқ қувват (80-100%)',
        'Ўрта қувват (51-79%)',
        'Паст қувват (0-50%)',
        'Вақтинча ишламаяпти',
        'Тугатилган',
        'Мавсумий',
        'Бошқа',
    ]

    if not get_group_post_monitoring_id(request):
        columns = [
            *columns_list
        ]
    else:
        columns = [
            "Ҳудудлар",
            *columns_list
        ]

    for col_num in range(len(columns)):
        column_letter = openpyxl.utils.get_column_letter(col_num + 1)
        ws.column_dimensions[column_letter].width = 25
        ws.cell(row=row_num, column=col_num + 1, value=columns[col_num])
        ws.cell(row=row_num, column=col_num + 1).style = cell_style

    form = TableSubtotalsForm()
    all_fields = form.fields.keys()
    fields = [
        *all_fields
    ]

    department_fields = [
        'district__district',
        *fields
    ]

    rows = filter_export_postmonitoring_tables(request, filter_slug, Subtotals, fields, department_fields)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.cell(row=row_num, column=col_num + 1, value=str(row[col_num]))

    wb.save(response)
    return response



def get_data_table(request, model_name, page_title):
    model = loiha_models.get(model_name, None)
    if not model:
        # Handle invalid model name
        raise ValueError("Invalid model name: %s" % model_name)

    if get_group_loiha_id(request):
        table_data = show_data_table_to_departament(model)
    else:
        table_data = show_data_table(request, model)

    page_obj = paginate_page(request, table_data)
    context = get_context_data(page_obj, page_title, projects_department, table_data)
    return render(request, src[model_name], context)


def get_data_table_Loiha41(request):
    return get_data_table(request, 'loiha41', 'Илова-4.1')


def get_data_table_Loiha52(request):
    return get_data_table(request, 'loiha52', 'Илова-5.2')


def get_data_table_Loiha14(request):
    return get_data_table(request, 'loiha14', 'Илова-14')


def get_data_table_Loiha131(request):
    return get_data_table(request, 'loiha131', 'Илова-13.1')


def get_data_table_Loiha122(request):
    return get_data_table(request, 'loiha122', 'Илова-12.2')


def get_data_table_Loiha121(request):
    return get_data_table(request, 'loiha121', 'Илова-12.1')


def get_data_table_Loiha12(request):
    return get_data_table(request, 'loiha12', 'Илова-12')


def get_data_table_Loiha10(request):
    return get_data_table(request, 'loiha10', 'Илова-10')


def get_data_table_Loiha6(request):
    return get_data_table(request, 'loiha6', 'Илова-6')


def get_data_table_Loiha13(request):
    return get_data_table(request, 'loiha13', 'Илова-13')


# All copied elements located in tests.py


def get_data_table_filter(request, model_name, page_title, filter_slug):
    model = loiha_models.get(model_name, None)
    if not model:
        # Handle invalid model name
        raise ValueError("Invalid model name: %s" % model_name)

    if not get_group_loiha_id(request):
        table_data = show_data_table(request, model)
    else:
        table_data = show_data_table_to_departament(model)

    table_data = filter_tables(filter_slug, table_data)

    page_obj = paginate_page(request, table_data)
    context = get_context_data(page_obj, page_title, projects_department, table_data)
    return render(request, src[model_name], context)


def table_filter_loiha41(request, filter_slug):
    return get_data_table_filter(request, 'loiha41', 'Илова-4.1', filter_slug)


def table_filter_loiha52(request, filter_slug):
    return get_data_table_filter(request, 'loiha52', 'Илова-5.2', filter_slug)


def table_filter_loiha14(request, filter_slug):
    return get_data_table_filter(request, 'loiha14', 'Илова-14', filter_slug)


def table_filter_loiha131(request, filter_slug):
    return get_data_table_filter(request, 'loiha131', 'Илова-13.1', filter_slug)


def table_filter_loiha122(request, filter_slug):
    return get_data_table_filter(request, 'loiha122', 'Илова-12.2', filter_slug)


def table_filter_loiha121(request, filter_slug):
    return get_data_table_filter(request, 'loiha121', 'Илова-12.1', filter_slug)


def table_filter_loiha12(request, filter_slug):
    return get_data_table_filter(request, 'loiha12', 'Илова-12', filter_slug)


def table_filter_loiha10(request, filter_slug):
    return get_data_table_filter(request, 'loiha10', 'Илова-10', filter_slug)


def table_filter_loiha6(request, filter_slug):
    return get_data_table_filter(request, 'loiha6', 'Илова-6', filter_slug)


def table_filter_loiha13(request, filter_slug):
    return get_data_table_filter(request, 'loiha13', 'Илова-13', filter_slug)


def logoutUser(request):
    logout(request)
    return render(request, 'mainapp/logout.html')


# TODO прочитать статью Django про итеграцию с телеграм ботом. Костамизация админкиз


@login_required
def add_data_table(request, model_name, model_form, redirect_url, page_title):
    model = loiha_models.get(model_name, None)
    table_data = show_data_table(request, model)
    page_obj = paginate_page(request, table_data)
    if request.method == 'POST':
        form = model_form(request.POST)
        if form.is_valid():
            try:
                model.objects.create(**form.cleaned_data, district=request.user.district)
                return redirect(redirect_url)
            except:
                form.add_error(None, 'Ошибка добавления данных')
    else:
        form = model_form()
    context = make_context_by_form(page_title, form, projects_department, page_obj)
    return render(request, f'mainapp/forms/form_{model_name}.html', context)


# @login_required
def add_data_table_Loiha41(request):
    return add_data_table(request, 'loiha41', TableFormLoiha41, 'table_loiha4.1', 'Илова-4.1')


def add_data_table_Loiha52(request):
    return add_data_table(request, 'loiha52', TableFormLoiha52, 'table_loiha5.2', 'Илова-5.2')


def add_data_table_Loiha14(request):
    return add_data_table(request, 'loiha14', TableFormLoiha14, 'table_loiha14', 'Илова-14')


def add_data_table_Loiha131(request):
    return add_data_table(request, 'loiha131', TableFormLoiha131, 'table_loiha13.1', 'Илова-13.1')


def add_data_table_Loiha122(request):
    return add_data_table(request, 'loiha122', TableFormLoiha122, 'table_loiha12.2', 'Илова-12.2')


def add_data_table_Loiha121(request):
    return add_data_table(request, 'loiha121', TableFormLoiha121, 'table_loiha12.1', 'Илова-12.1')


def add_data_table_Loiha12(request):
    return add_data_table(request, 'loiha12', TableFormLoiha12, 'table_loiha12', 'Илова-12')


def add_data_table_Loiha10(request):
    return add_data_table(request, 'loiha10', TableFormLoiha10, 'table_loiha10', 'Илова-10')


def add_data_table_Loiha6(request):
    return add_data_table(request, 'loiha6', TableFormLoiha6, 'table_loiha6', 'Илова-6')


def add_data_table_Loiha13(request):
    return add_data_table(request, 'loiha13', TableFormLoiha13, 'table_loiha13', 'Илова-13')


# Второй отдел

def get_data_table_2(request, model_name, page_title):
    model = second_department_models_dict.get(model_name, None)
    if not model:
        # Handle invalid model name
        raise ValueError("Invalid model name: %s" % model_name)

    if get_group_export_id(request):
        table_data = show_data_table_to_departament(model)
    else:
        table_data = show_data_table(request, model)

    page_obj = paginate_page(request, table_data)
    context = get_context_data(page_obj, page_title, second_department_tables_menu, table_data)
    return render(request, src[model_name], context)


def get_data_table_sanoat(request):
    return get_data_table_2(request, 'sanoat', 'Sanoat')


def get_data_table_kh(request):
    return get_data_table_2(request, 'kx', 'KX')


def get_data_table_first_table(request):
    return get_data_table_2(request, 'table_1', 'Таблица 1')


def get_data_table_kunliu(request):
    return get_data_table_2(request, 'kunliu', 'кунлиу')


def get_data_table_filter_second_departament(request, model_name, page_title, filter_slug):
    model = second_department_models_dict.get(model_name, None)
    if not model:
        # Handle invalid model name
        raise ValueError("Invalid model name: %s" % model_name)

    if not get_group_export_id(request):
        table_data = show_data_table(request, model)
    else:
        table_data = show_data_table_to_departament(model)

    table_data = filter_tables(filter_slug, table_data)

    page_obj = paginate_page(request, table_data)
    context = get_context_data(page_obj, page_title, second_department_tables_menu, table_data)
    return render(request, src[model_name], context)


def table_filter_sanoat(request, filter_slug):
    return get_data_table_filter_second_departament(request, 'sanoat', 'Sanoat', filter_slug)


def table_filter_kx(request, filter_slug):
    return get_data_table_filter_second_departament(request, 'kx', 'KX', filter_slug)


def table_filter_first_table(request, filter_slug):
    return get_data_table_filter_second_departament(request, 'table_1', 'Таблица 1', filter_slug)


def table_filter_kunliu(request, filter_slug):
    return get_data_table_filter_second_departament(request, 'kunliu', 'кунлиу', filter_slug)


@login_required
def add_data_table_second_department(request, model_name, model_form, redirect_url, page_title):
    model = second_department_models_dict.get(model_name, None)
    table_data = show_data_table(request, model)
    page_obj = paginate_page(request, table_data)
    if request.method == 'POST':
        form = model_form(request.POST)
        if form.is_valid():
            try:
                model.objects.create(**form.cleaned_data, district=request.user.district)
                return redirect(redirect_url)
            except:
                form.add_error(None, 'Ошибка добавления данных')
    else:
        form = model_form()
    context = make_context_by_form(page_title, form, second_department_tables_menu, page_obj)
    return render(request, f'mainapp/forms/second_departament/form_{model_name}.html', context)


def add_data_table_sanoat(request):
    return add_data_table_second_department(request, 'sanoat', TableFormSanoat, 'table_sanoat', 'Sanoat')


def add_data_table_kx(request):
    return add_data_table_second_department(request, 'kx', TableFormKX, 'table_kx', 'KX')


def add_data_table_1(request):
    return add_data_table_second_department(request, 'table_1', TableFirstForm, 'table_first', 'Таблица 1')


def add_data_table_kunliu(request):
    return add_data_table_second_department(request, 'kunliu', TableKunliuForm, 'table_kunliu', 'кунлиу')


# 3 Сводный отдел
def get_data_table_3(request, model_name, page_title):
    model = third_department_models_dict.get(model_name, None)
    if not model:
        # Handle invalid model name
        raise ValueError("Invalid model name: %s" % model_name)

    if get_group_vault_id(request):
        table_data = show_data_table_to_departament(model)
    else:
        table_data = show_data_table(request, model)

    page_obj = paginate_page(request, table_data)
    context = get_context_data(page_obj, page_title, third_department_tables_menu, table_data)
    return render(request, src[model_name], context)


def get_data_table_jami(request):
    return get_data_table_3(request, 'jami', 'Жами свод')


def get_data_table_quarter(request):
    return get_data_table_3(request, 'quarter', 'Cвод чорак')


def get_data_table_month(request):
    return get_data_table_3(request, 'monthly', 'Cвод ойлар')


def get_data_table_bank(request):
    return get_data_table_3(request, 'bank', 'Cвод банк')


def get_data_table_reja(request):
    return get_data_table_3(request, 'reja', 'Cвод режа')


def get_data_table_tarmok(request):
    return get_data_table_3(request, 'tarmok', 'Cвод тармок')


# Жами Свод


def get_data_table_filter_third_departament(request, model_name, page_title, filter_slug):
    model = third_department_models_dict.get(model_name, None)
    if not model:
        # Handle invalid model name
        raise ValueError("Invalid model name: %s" % model_name)

    if not get_group_vault_id(request):
        table_data = show_data_table(request, model)
    else:
        table_data = show_data_table_to_departament(model)

    table_data = filter_tables(filter_slug, table_data)
    page_obj = paginate_page(request, table_data)
    context = get_context_data(page_obj, page_title, third_department_tables_menu, table_data)
    return render(request, src[model_name], context)


def table_filter_table_jami(request, filter_slug):
    return get_data_table_filter_third_departament(request, 'jami', 'Жами свод', filter_slug)


def table_filter_table_quarter(request, filter_slug):
    return get_data_table_filter_third_departament(request, 'quarter', 'Cвод чорак', filter_slug)


def table_filter_table_month(request, filter_slug):
    return get_data_table_filter_third_departament(request, 'monthly', 'Cвод ойлар', filter_slug)


def table_filter_table_bank(request, filter_slug):
    return get_data_table_filter_third_departament(request, 'bank', 'Cвод банк', filter_slug)


def table_filter_table_reja(request, filter_slug):
    return get_data_table_filter_third_departament(request, 'reja', 'Cвод режа', filter_slug)


def table_filter_table_tarmok(request, filter_slug):
    return get_data_table_filter_third_departament(request, 'tarmok', 'Cвод тармок', filter_slug)


@login_required
def add_data_table_third_department(request, model_name, model_form, redirect_url, page_title):
    model = third_department_models_dict.get(model_name, None)
    table_data = show_data_table(request, model)
    page_obj = paginate_page(request, table_data)
    if request.method == 'POST':
        form = model_form(request.POST)
        if form.is_valid():
            try:
                model.objects.create(**form.cleaned_data, district=request.user.district)
                return redirect(redirect_url)
            except:
                form.add_error(None, 'Ошибка добавления данных')
    else:
        form = model_form()
    context = make_context_by_form(page_title, form, third_department_tables_menu, page_obj)
    return render(request, f'mainapp/forms/third_departament/form_{model_name}.html', context)


def add_data_table_jami(request):
    return add_data_table_third_department(request, 'jami', TableJamiForm, 'table_jami', 'Жами свод')


def add_data_table_quarter(request):
    return add_data_table_third_department(request, 'quarter', TableQuarterForm, 'table_quarter', 'Cвод чорак')


def add_data_table_month(request):
    return add_data_table_third_department(request, 'monthly', TableMonthForm, 'table_monthly', 'Cвод ойлар')


def add_data_table_bank(request):
    return add_data_table_third_department(request, 'bank', TableBankForm, 'table_bank', 'Cвод банк')


def add_data_table_reja(request):
    return add_data_table_third_department(request, 'reja', TableRejaForm, 'table_reja', 'Cвод режа')


def add_data_table_tarmok(request):
    return add_data_table_third_department(request, 'tarmok', TableTarmokForm, 'table_tarmok', 'Cвод тармок')


def get_data_table_4(request, model_name, page_title):
    model = fourth_department_models_dict.get(model_name, None)
    if not model:
        # Handle invalid model name
        raise ValueError("Invalid model name: %s" % model_name)

    if get_group_post_monitoring_id(request):
        table_data = show_data_table_to_departament(model)
    else:
        table_data = show_data_table(request, model)

    page_obj = paginate_page(request, table_data)
    context = get_context_data(page_obj, page_title, fourth_department_tables_menu, table_data)
    return render(request, src[model_name], context)


def get_data_table_manzil(request):
    return get_data_table_4(request, 'manzil', '2017-2020 манзилли')


def get_data_table_subtotals(request):
    return get_data_table_4(request, 'subtotals', 'Промежуточный итоги')


def get_data_table_addressed(request):
    return get_data_table_4(request, 'addressed', 'Манзилли')


def get_data_table_networkAdministrations(request):
    return get_data_table_4(request, 'networkAdministrations', 'Тармоқ бошқармалари')


def get_data_table_totalCleaning(request):
    return get_data_table_4(request, 'totalCleaning', 'ЖАМИ чистка')


def get_data_table_totalCleaningNetwork(request):
    return get_data_table_4(request, 'totalCleaningNetwork', 'ЖАМИ чистка тармоқ')


def get_data_table_totalDone(request):
    return get_data_table_4(request, 'totalDone', 'ЖАМИ бажарилган')


def get_data_table_totalCompletedNetwork(request):
    return get_data_table_4(request, 'totalCompletedNetwork', 'ЖАМИ бажарилган тармоқ')


def get_data_table_totalProblem(request):
    return get_data_table_4(request, 'totalProblem', 'ЖАМИ муаммо')


def get_data_table_performanceAddressed(request):
    return get_data_table_4(request, 'performanceAddressed', 'манзилли')


def get_data_table_filter_fourth_departament(request, model_name, page_title, filter_slug):
    model = fourth_department_models_dict.get(model_name, None)
    if not model:
        # Handle invalid model name
        raise ValueError("Invalid model name: %s" % model_name)

    if not get_group_post_monitoring_id(request):
        table_data = show_data_table(request, model)
    else:
        table_data = show_data_table_to_departament(model)

    table_data = filter_tables(filter_slug, table_data)
    page_obj = paginate_page(request, table_data)
    context = get_context_data(page_obj, page_title, fourth_department_tables_menu, table_data)
    return render(request, src[model_name], context)


def table_filter_table_manzil(request, filter_slug):
    return get_data_table_filter_fourth_departament(request, 'manzil', '2017-2020 манзилли', filter_slug)


def table_filter_table_subtotals(request, filter_slug):
    return get_data_table_filter_fourth_departament(request, 'subtotals', 'Промежуточный итоги', filter_slug)


def table_filter_table_addressed(request, filter_slug):
    return get_data_table_filter_fourth_departament(request, 'addressed', 'Манзилли', filter_slug)


def table_filter_table_networkAdministrations(request, filter_slug):
    return get_data_table_filter_fourth_departament(request, 'networkAdministrations', 'Тармоқ бошқармалари',
                                                    filter_slug)


def table_filter_table_totalCleaning(request, filter_slug):
    return get_data_table_filter_fourth_departament(request, 'totalCleaning', 'ЖАМИ чистка', filter_slug)


def table_filter_table_totalCleaningNetwork(request, filter_slug):
    return get_data_table_filter_fourth_departament(request, 'totalCleaningNetwork', 'ЖАМИ чистка тармоқ', filter_slug)


def table_filter_table_totalDone(request, filter_slug):
    return get_data_table_filter_fourth_departament(request, 'totalDone', 'ЖАМИ бажарилган', filter_slug)


def table_filter_table_totalCompletedNetwork(request, filter_slug):
    return get_data_table_filter_fourth_departament(request, 'totalCompletedNetwork', 'ЖАМИ бажарилган тармоқ',
                                                    filter_slug)


def table_filter_table_totalProblem(request, filter_slug):
    return get_data_table_filter_fourth_departament(request, 'totalProblem', 'ЖАМИ муаммо', filter_slug)


def table_filter_table_performanceAddressed(request, filter_slug):
    return get_data_table_filter_fourth_departament(request, 'performanceAddressed', 'манзилли', filter_slug)


@login_required
def add_data_table_fourth_department(request, model_name, model_form, redirect_url, page_title):
    model = fourth_department_models_dict.get(model_name, None)
    table_data = show_data_table(request, model)
    page_obj = paginate_page(request, table_data)
    if request.method == 'POST':
        form = model_form(request.POST)
        if form.is_valid():
            try:
                model.objects.create(**form.cleaned_data, district=request.user.district)
                return redirect(redirect_url)
            except:
                form.add_error(None, 'Ошибка добавления данных')
    else:
        form = model_form()
    context = make_context_by_form(page_title, form, fourth_department_tables_menu, page_obj)
    return render(request, f'mainapp/forms/fourth_department/form_{model_name}.html', context)


def add_data_table_manzil(request):
    return add_data_table_fourth_department(request, 'manzil', TableManzilForm, 'table_manzil', '2017-2020 манзилли')


def add_data_table_subtotals(request):
    return add_data_table_fourth_department(request, 'subtotals', TableSubtotalsForm, 'table_subtotals',
                                            'Промежуточный итоги')


def add_data_table_addressed(request):
    return add_data_table_fourth_department(request, 'addressed', TableAddressedForm, 'table_addressed', 'Манзилли')


def add_data_table_networkAdministrations(request):
    return add_data_table_fourth_department(request, 'networkAdministrations', TableNetworkAdministrationsForm,
                                            'table_networkAdministrations', 'Тармоқ бошқармалари')


def add_data_table_totalCleaning(request):
    return add_data_table_fourth_department(request, 'totalCleaning', TableTotalCleaningForm, 'table_totalCleaning',
                                            'ЖАМИ чистка')


def add_data_table_totalCleaningNetwork(request):
    return add_data_table_fourth_department(request, 'totalCleaningNetwork', TableTotalCleaningNetworkForm,
                                            'table_totalCleaningNetwork', 'ЖАМИ чистка тармоқ')


def add_data_table_totalDone(request):
    return add_data_table_fourth_department(request, 'totalDone', TableTotalDoneForm, 'table_totalDone',
                                            'ЖАМИ бажарилган')


def add_data_table_totalCompletedNetwork(request):
    return add_data_table_fourth_department(request, 'totalCompletedNetwork', TableTotalCompletedNetworkForm,
                                            'table_totalCompletedNetwork', 'ЖАМИ бажарилган тармоқ')


def add_data_table_totalProblem(request):
    return add_data_table_fourth_department(request, 'totalProblem', TableTotalProblemForm, 'table_totalProblem',
                                            'ЖАМИ муаммо')


def add_data_table_performanceAddressed(request):
    return add_data_table_fourth_department(request, 'performanceAddressed', TablePerformanceIsAddressedForm,
                                            'table_performanceAddressed', 'манзилли')
